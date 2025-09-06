import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from datetime import timedelta

def generate_full_attendance_excel(students, timetable_entries, start_date, end_date, batch_name, phase, faculty_name, subject_name):
    lecture_map = {
        "08:45-09:45": "Lec 1",
        "09:45-10:45": "Lec 2",
        "11:30-12:30": "Lec 3",
        "12:30-13:30": "Lec 4",
        "13:45-14:45": "Lec 5"
    }

    date_range = pd.date_range(start=start_date, end=end_date)
    att_df = pd.DataFrame([{
        "Sr No": i + 1,
        "Enrollment No": s.enrollment_no,
        "Roll No": s.roll_no,
        "Student Name": s.name
    } for i, s in enumerate(students)])

    col_headers = []
    week_start = start_date
    week_num = 1

    for d in date_range:
        if (d - week_start).days > 6:
            week_num += 1
            week_start = d
        weekday = d.strftime("%A")
        for e in timetable_entries:
            if e.day == weekday and e.time in lecture_map:
                lec = lecture_map[e.time]
                col = f"{phase}|T{phase[-1]}_W{week_num}|{d.strftime('%d-%b')}|{lec}"
                att_df[col] = "P"
                col_headers.append((phase, f"T{phase[-1]}_W{week_num}", d.strftime('%d-%b'), lec))

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        att_df.to_excel(writer, index=False, startrow=6, sheet_name="Attendance")
        ws = writer.sheets["Attendance"]

        for i, (p, w, d, lec) in enumerate(col_headers):
            col_letter = get_column_letter(i + 5)
            ws[f"{col_letter}2"] = p
            ws[f"{col_letter}3"] = w
            ws[f"{col_letter}4"] = d
            ws[f"{col_letter}5"] = lec
            for r in range(2, 6):
                ws[f"{col_letter}{r}"].alignment = Alignment(horizontal='center')
                ws[f"{col_letter}{r}"].font = Font(bold=True)

        def merge_headers(headers, row):
            last = None
            start = 0
            for i, val in enumerate(headers + [None]):
                if val != last:
                    if last is not None:
                        c1 = get_column_letter(start + 5)
                        c2 = get_column_letter(i + 4)
                        if c1 != c2:
                            ws.merge_cells(f"{c1}{row}:{c2}{row}")
                    start = i
                    last = val

        merge_headers([h[0] for h in col_headers], 2)
        merge_headers([h[1] for h in col_headers], 3)
        merge_headers([h[2] for h in col_headers], 4)

        ws["A1"] = f"Subject: {subject_name}"
        ws["B1"] = f"Faculty: {faculty_name}"
        ws["C1"] = f"Batch: {batch_name}"

        thin = Side(border_style="thin", color="000000")
        for r in range(2, att_df.shape[0] + 7):
            for c in range(1, len(att_df.columns) + 1):
                ws.cell(row=r, column=c).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    output.seek(0)
    return output
