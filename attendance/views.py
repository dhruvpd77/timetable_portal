from django.shortcuts import render, redirect, HttpResponse
from django.contrib import messages
from core.models import Department, Batch, Timetable, TimetableEntry, Faculty, CourseSpec
from .models import AttendanceSessionWindow, Student

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font
from datetime import datetime, timedelta
from io import BytesIO
from collections import defaultdict

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def safe_str(val):
    if pd.isnull(val):
        return ""
    return str(val).strip()
def clean_number(val):
    if pd.isnull(val):
        return ""
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    return s
import pandas as pd
from django.contrib import messages
import os

def manage_students(request):
    user_department = Department.objects.filter(user=request.user).first()
    batches = Batch.objects.filter(department=user_department)

    selected_batch_id = request.GET.get('batch') or request.POST.get('batch')
    selected_batch = None
    if selected_batch_id:
        selected_batch = Batch.objects.filter(id=selected_batch_id, department=user_department).first()

    students = Student.objects.filter(department=user_department, batch=selected_batch) if selected_batch else []

    if request.method == "POST" and 'upload' in request.POST:
        file = request.FILES.get('csv_file')
        if not file:
            messages.error(request, "Please upload a CSV or Excel file.")
        else:
            filename = file.name.lower()
            try:
                # Decide which reader to use
                if filename.endswith('.csv'):
                    df = pd.read_csv(file)
                elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                    df = pd.read_excel(file)
                else:
                    messages.error(request, "Unsupported file format. Please upload a CSV, XLSX, or XLS file.")
                    return redirect(request.path + (f"?batch={selected_batch.id}" if selected_batch else ""))

                df.columns = [col.strip().lower() for col in df.columns]
                required = ['name', 'enrollment no', 'roll no']
                if not all(col in df.columns for col in required):
                    messages.error(request, "File must contain columns: Name, Enrollment No, Roll No")
                else:
                    df = df.rename(columns={
                        'name': 'name',
                        'enrollment no': 'enrollment_no',
                        'roll no': 'roll_no'
                    })
                    valid_rows = df.dropna(subset=['name', 'enrollment_no', 'roll_no'])
                    objs = []
                    for _, row in valid_rows.iterrows():
                        objs.append(Student(
                            department=user_department,
                            batch=selected_batch,
                            name=safe_str(row['name']),
                            enrollment_no=clean_number(row['enrollment_no']),
                            roll_no=clean_number(row['roll_no'])
                        ))
                    Student.objects.filter(department=user_department, batch=selected_batch).delete()
                    Student.objects.bulk_create(objs)
                    if len(valid_rows) < len(df):
                        messages.warning(request, f"{len(df) - len(valid_rows)} row(s) with missing data were skipped.")
                    messages.success(request, "Student data uploaded successfully!")
                    return redirect(request.path + f"?batch={selected_batch.id}")

            except Exception as e:
                messages.error(request, f"Failed to read file: {str(e)}")

    context = {
        "department": user_department,
        "batches": batches,
        "selected_batch": selected_batch,
        "students": students
    }
    return render(request, "attendance/manage_students.html", context)


def get_dates_in_window(start, end, weekday):
    if start is None or end is None:
        return []
    result = []
    current = start
    while current <= end:
        if current.weekday() == weekday:
            result.append(current)
        current += timedelta(days=1)
    return result


def chunk_dates_into_weeks(dates):
    weeks = []
    if not dates:
        return weeks
    week = []
    current_iso_week = dates[0].isocalendar()[1]
    for dt in dates:
        iso_week = dt.isocalendar()[1]
        if iso_week != current_iso_week:
            weeks.append(week)
            week = [dt]
            current_iso_week = iso_week
        else:
            week.append(dt)
    if week:
        weeks.append(week)
    return weeks


def attendance_sheet_generator(request):
    # *** FIX: Define these at start so they are always available ***
    user_department = Department.objects.filter(user=request.user).first()
    session_window = AttendanceSessionWindow.objects.filter(department=user_department).first()
    batches = Batch.objects.filter(department=user_department)
    timetables = Timetable.objects.filter(department=user_department)
    msg = None

    if request.method == "POST" and "save_windows" in request.POST:
        t1_start = request.POST.get("t1_start")
        t1_end = request.POST.get("t1_end")
        t2_start = request.POST.get("t2_start")
        t2_end = request.POST.get("t2_end")
        t3_start = request.POST.get("t3_start")
        t3_end = request.POST.get("t3_end")
        t4_start = request.POST.get("t4_start")
        t4_end = request.POST.get("t4_end")
        AttendanceSessionWindow.objects.update_or_create(
            department=user_department,
            defaults=dict(
                t1_start=t1_start, t1_end=t1_end,
                t2_start=t2_start, t2_end=t2_end,
                t3_start=t3_start, t3_end=t3_end,
                t4_start=t4_start, t4_end=t4_end
            )
        )
        session_window = AttendanceSessionWindow.objects.get(department=user_department)
        msg = "Session windows saved successfully!"

    if request.method == "POST" and "download_attendance" in request.POST:
        batch_id = request.POST.get("batch")
        timetable_id = request.POST.get("timetable")
        if not batch_id or not timetable_id or not session_window:
            messages.error(request, "All fields must be selected and session window set up first!")
            return redirect(request.path)
        batch = Batch.objects.get(id=batch_id)
        timetable = Timetable.objects.get(id=timetable_id)
        students = Student.objects.filter(department=user_department, batch=batch)
        entries = TimetableEntry.objects.filter(
            department=user_department, batch=batch, timetable=timetable
        )

        wb = Workbook()
        wb.remove(wb.active)

        win_list = [
            ("T1", session_window.t1_start, session_window.t1_end),
            ("T2", session_window.t2_start, session_window.t2_end),
            ("T3", session_window.t3_start, session_window.t3_end),
            ("T4", session_window.t4_start, session_window.t4_end),
        ]

        subj_faculty_dict = defaultdict(list)
        for e in entries:
            key = (e.subject.subject_name, e.faculty.short_name)
            subj_faculty_dict[key].append(e)

        for (subj, faculty_short), entry_list in subj_faculty_dict.items():
            ws_name = f"{subj}-{faculty_short}"[:31]
            ws = wb.create_sheet(title=ws_name)

            fixed_cols = 3
            # Headers for student info, spanning rows 4 to 7
            for col, title in enumerate(["Roll No", "Name", "Enrollment No"], start=1):
                ws.cell(row=4, column=col, value=title)
                ws.merge_cells(start_row=4, start_column=col, end_row=7, end_column=col)
                c = ws.cell(row=4, column=col)
                c.alignment = Alignment(vertical='center', horizontal='center')
                c.font = Font(bold=True)
                c.border = thin_border

            current_col = fixed_cols + 1
            for term_label, t_start, t_end in win_list:
                if not (t_start and t_end):
                    continue
                day_lecture_times = defaultdict(list)
                for en in entry_list:
                    try:
                        weekday_idx = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday'].index(en.day.lower())
                    except:
                        continue
                    day_lecture_times[weekday_idx].append(en.time)
                for k in day_lecture_times:
                    day_lecture_times[k] = sorted(set(day_lecture_times[k]))

                term_dates = []
                for day_idx in day_lecture_times.keys():
                    term_dates += get_dates_in_window(t_start, t_end, day_idx)
                term_dates = sorted(term_dates)

                weeks = chunk_dates_into_weeks(term_dates)
                week_col_counts = []
                week_date_lectures = []

                for wk in weeks:
                    col_count = 0
                    date_lecture_list = []
                    for dt in wk:
                        day_idx = dt.weekday()
                        lectures = day_lecture_times.get(day_idx, [])
                        date_lecture_list.append((dt, lectures))
                        col_count += len(lectures) if lectures else 1
                    week_col_counts.append(col_count)
                    week_date_lectures.append(date_lecture_list)

                total_term_cols = sum(week_col_counts)
                if total_term_cols == 0:
                    continue

                ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + total_term_cols - 1)
                c_term = ws.cell(row=1, column=current_col, value=term_label)
                c_term.alignment = Alignment(horizontal='center', vertical='center')
                c_term.font = Font(bold=True)
                for col_idx in range(current_col, current_col + total_term_cols):
                    ws.cell(row=1, column=col_idx).border = thin_border
                    ws.cell(row=2, column=col_idx).border = thin_border

                col_pointer = current_col
                for idx_wk, week_count in enumerate(week_col_counts):
                    if week_count == 0:
                        continue
                    ws.merge_cells(start_row=2, start_column=col_pointer, end_row=2, end_column=col_pointer + week_count - 1)
                    c_week = ws.cell(row=2, column=col_pointer, value=f"Week {idx_wk + 1}")
                    c_week.alignment = Alignment(horizontal='center', vertical='center')
                    c_week.font = Font(bold=True)
                    for col_idx in range(col_pointer, col_pointer + week_count):
                        ws.cell(row=2, column=col_idx).border = thin_border

                    date_lecture_list = week_date_lectures[idx_wk]
                    date_col_start = col_pointer
                    for dt, lectures in date_lecture_list:
                        lecture_count = len(lectures) if lectures else 1
                        if lecture_count > 1:
                            ws.merge_cells(start_row=3, start_column=date_col_start, end_row=3, end_column=date_col_start + lecture_count - 1)

                        c_date = ws.cell(row=3, column=date_col_start, value=dt.strftime("%d-%b-%Y"))
                        c_date.alignment = Alignment(horizontal='center', vertical='center')
                        c_date.font = Font(bold=True)
                        for col_idx in range(date_col_start, date_col_start + lecture_count):
                            ws.cell(row=3, column=col_idx).border = thin_border

                        if lectures:
                            for idx_lec, lec_time in enumerate(lectures):
                                c_lec = ws.cell(row=4, column=date_col_start + idx_lec, value=f"Lecture {idx_lec + 1}\n{lec_time}")
                                c_lec.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                                c_lec.font = Font(italic=True)
                                c_lec.border = thin_border
                        else:
                            c_lec = ws.cell(row=4, column=date_col_start, value="Lecture 1")
                            c_lec.alignment = Alignment(horizontal='center', vertical='center')
                            c_lec.font = Font(italic=True)
                            c_lec.border = thin_border

                        date_col_start += lecture_count
                    col_pointer += week_count

                current_col += total_term_cols

            start_data_row = 8
            for idx_stu, stu in enumerate(students):
                ws.cell(row=start_data_row + idx_stu, column=1, value=stu.roll_no).border = thin_border
                ws.cell(row=start_data_row + idx_stu, column=2, value=stu.name).border = thin_border
                ws.cell(row=start_data_row + idx_stu, column=3, value=stu.enrollment_no).border = thin_border
                for c in range(fixed_cols + 1, current_col):
                    ws.cell(row=start_data_row + idx_stu, column=c, value="").border = thin_border

            for col in range(1, fixed_cols + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            for col in range(fixed_cols + 1, current_col):
                ws.column_dimensions[get_column_letter(col)].width = 12

            ws.freeze_panes = ws['D8']

        memfile = BytesIO()
        wb.save(memfile)
        memfile.seek(0)
        filename = f"{batch.name}_attendance_sheet.xlsx"
        return HttpResponse(
            memfile.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )

    context = {
        "batches": batches,
        "timetables": timetables,
        "session_window": session_window,
        "msg": msg,
    }
    return render(request,"attendance/attendance_sheet_generator.html",context)
# attendance/views.py






import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from io import BytesIO

from core.models import Department, Faculty, TimetableEntry, Batch
from attendance.models import Student, AttendanceSessionWindow
from faculty_attendance.models import FacultyAttendance


from collections import defaultdict
from datetime import datetime, timedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render

@login_required
def daily_absent(request):
    # Get user's department via Department.user
    department = Department.objects.filter(user=request.user).first()
    if not department:
        messages.error(request, "Your user is not assigned to any department.")
        return redirect('home')  # Adjust appropriately

    # Get attendance session window for this department
    session_window = AttendanceSessionWindow.objects.filter(department=department).first()

    def get_valid_dates(dept, session):
        if not session:
            return []
        lecture_days = set(
            TimetableEntry.objects.filter(department=dept)
            .values_list("day", flat=True).distinct()
        )
        lecture_days = {day.lower() for day in lecture_days if day}
        valid_dates = []
        for start, end in [
            (session.t1_start, session.t1_end),
            (session.t2_start, session.t2_end),
            (session.t3_start, session.t3_end),
            (session.t4_start, session.t4_end),
        ]:
            if not start or not end:
                continue
            current = start
            while current <= end:
                if current.strftime("%A").lower() in lecture_days:
                    valid_dates.append(current)
                current += timedelta(days=1)
        return sorted(valid_dates)

    valid_dates = get_valid_dates(department, session_window)

    selected_date_str = request.GET.get("date") or request.POST.get("date")
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, "%Y-%m-%d").date()
            if selected_date not in valid_dates:
                selected_date = valid_dates[0] if valid_dates else None
        except Exception:
            selected_date = valid_dates[0] if valid_dates else None
    else:
        selected_date = valid_dates[0] if valid_dates else None

    weekday_name = selected_date.strftime("%A") if selected_date else None
    lectures = []
    if weekday_name:
        lectures = TimetableEntry.objects.filter(
            department=department,
            day__iexact=weekday_name,
        ).select_related("batch", "faculty", "subject").order_by("batch__name", "time")

    # Group lectures by batch (using batch name as key)
    lectures_by_batch = defaultdict(list)
    for lec in lectures:
        lectures_by_batch[lec.batch.name].append(lec)

    # Load students per batch
    students_per_batch = {}
    from core.models import Batch
    for batch_name in lectures_by_batch:
        batch_obj = Batch.objects.filter(name=batch_name, department=department).first()
        if batch_obj:
            students_per_batch[batch_obj.id] = list(
                Student.objects.filter(department=department, batch=batch_obj.name)
                .values("roll_no", "name")
            )

    # Load existing attendance for this date
    attendance_qs = FacultyAttendance.objects.filter(
        faculty__department=department, date=selected_date
    )
    attendance_map = defaultdict(lambda: defaultdict(list))
    for att in attendance_qs:
        absent_list = att.absent_roll_numbers.split(",") if att.absent_roll_numbers else []
        attendance_map[att.batch.id][att.lecture_number] = absent_list

    # Determine missing attendance records
    missing_entries = []
    for lec in lectures:
        att = FacultyAttendance.objects.filter(
            faculty=lec.faculty,
            batch=lec.batch,
            date=selected_date,
            lecture_number=lec.time,
        ).first()
        if not att or not att.absent_roll_numbers:
            missing_entries.append((lec.faculty.full_name, lec.batch.name, lec.time))

    can_generate_report = not missing_entries and bool(lectures)

    if request.method == "POST":
        # Removed save_attendance logic entirely

        if "generate_report" in request.POST:
            if missing_entries:
                messages.error(request, "Cannot generate report: attendance missing for some lectures.")
                return redirect(f"{request.path}?date={selected_date}")
            return generate_excel_response(selected_date, department, lectures_by_batch)

    return render(
        request,
        "attendance/daily_absent_single_page.html",
        {
            "valid_dates": valid_dates,
            "selected_date": selected_date,
            "lectures_by_batch": lectures_by_batch,
            "students_per_batch": students_per_batch,
            "attendance_map": attendance_map,
            "missing_entries": missing_entries,
            "can_generate_report": can_generate_report,
        },
    )

def generate_excel_response(selected_date, department, lectures_by_batch):
    import math

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Absent"

    title = f"Daily Absent Report - {department.name} - {selected_date.strftime('%d-%m-%Y')}"
    max_batches_per_row = 2
    headers = ["No", "Subject", "Faculty", "Absent Nos"]
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    batch_names = sorted(lectures_by_batch.keys())
    pairs = [batch_names[i:i+max_batches_per_row] for i in range(0, len(batch_names), max_batches_per_row)]
    current_row = 1

    # Report Title (merge all relevant columns)
    n_cols = (len(pairs[0]) * 5) if pairs else 5
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=n_cols)
    cell = ws.cell(row=current_row, column=1, value=title)
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    current_row += 2

    for pair in pairs:
        # Batch headers
        col = 1
        for batch in pair:
            ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row, end_column=col+3)
            batch_cell = ws.cell(row=current_row, column=col, value=f"Batch: {batch}")
            batch_cell.font = Font(bold=True, color="FFFFFF")
            batch_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            batch_cell.alignment = Alignment(horizontal='center')
            # Leave header cell for blank column
            col += 5
        current_row += 1
        # Column headers + blank col at end
        col = 1
        for batch in pair:
            for cidx, header in enumerate(headers):
                cell = ws.cell(row=current_row, column=col + cidx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            # blank column
            col += 5
        current_row += 1
        # Gather lectures, "wrap" Absent Nos as 5-per row
        max_lects = max(len(lectures_by_batch[batch]) for batch in pair)
        col = 1
        to_continue = []
        for batch in pair:
            lec_list = lectures_by_batch[batch]
            # Build "rows" for this batch
            rows = []
            for idx, lec in enumerate(lec_list, 1):
                att = FacultyAttendance.objects.filter(
                    faculty=lec.faculty,
                    batch=lec.batch,
                    date=selected_date,
                    lecture_number=lec.time
                ).first()
                absent_nos = att.absent_roll_numbers if att else ""
                absents = [a.strip() for a in absent_nos.split(',')] if absent_nos else []
                # chunk absent numbers 5 per row
                if not absents:
                    rows.append([idx, lec.subject.subject_name, lec.faculty.full_name, ''])
                else:
                    for i in range(0, len(absents), 5):
                        if i == 0:
                            rows.append([
                                idx, lec.subject.subject_name, lec.faculty.full_name, ', '.join(absents[i:i+5])
                            ])
                        else:
                            rows.append(['', '', '', ', '.join(absents[i:i+5])])
            to_continue.append(rows)
        # Pad each batch to same height
        block_height = max(len(x) for x in to_continue)
        for idx, data_rows in enumerate(to_continue):
            while len(data_rows) < block_height:
                data_rows.append(['', '', '', ''])
        # Write rows, two batch-blocks per line
        for ridx in range(block_height):
            col = 1
            for data_rows in to_continue:
                for cidx, value in enumerate(data_rows[ridx]):
                    cell = ws.cell(row=current_row, column=col + cidx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center' if cidx == 0 else 'left')
                col += 5
            current_row += 1
        # Extra blank line between each batch-pair
        current_row += 1

    # Set column widths for up to 2 batches/row (adjust as needed)
    for b in range(2):
        offset = b * 5
        ws.column_dimensions[get_column_letter(1+offset)].width = 5
        ws.column_dimensions[get_column_letter(2+offset)].width = 28
        ws.column_dimensions[get_column_letter(3+offset)].width = 22
        ws.column_dimensions[get_column_letter(4+offset)].width = 35

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"DailyAbsent_{selected_date.strftime('%Y-%m-%d')}.xlsx"
    return HttpResponse(output,
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="{filename}"'})
import os
from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO
from urllib.parse import quote

from django.conf import settings
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.contrib.auth.decorators import login_required

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from core.models import Department, Batch, Timetable, TimetableEntry
from attendance.models import Student, AttendanceSessionWindow
from faculty_attendance.models import FacultyAttendance
from openpyxl.styles import Font, PatternFill

def chunk_list_by_weeks(dates):
    weeks = []
    week = []
    last_week_num = None
    for dt in sorted(dates):
        wk_num = dt.isocalendar()[1]
        if wk_num != last_week_num and week:
            weeks.append(week)
            week = []
        week.append(dt)
        last_week_num = wk_num
    if week:
        weeks.append(week)
    return weeks


@login_required
def attendance_sheet_manager(request):
    user = request.user
    user_department = Department.objects.filter(user=user).first()
    if not user_department:
        messages.error(request, "You are not assigned to any department.")
        return redirect("home")

    session_window = AttendanceSessionWindow.objects.filter(department=user_department).first()
    if not session_window:
        messages.error(request, "Please set up attendance sessions first.")
        return redirect("attendance_session_setup")

    timetables = Timetable.objects.filter(department=user_department)
    phases = ["T1", "T2", "T3", "T4"]

    lecture_days = {d.lower() for d in TimetableEntry.objects.filter(department=user_department).values_list('day', flat=True) if d}

    week_blocks = []
    phase_week_to_continuous = {}
    continuous_week_counter = 1
    week_map = {}

    for phase in phases:
        start_date = getattr(session_window, f"{phase.lower()}_start", None)
        end_date = getattr(session_window, f"{phase.lower()}_end", None)
        if not start_date or not end_date:
            week_map[phase] = []
            continue

        valid_dates = []
        current = start_date
        while current <= end_date:
            if current.strftime("%A").lower() in lecture_days:
                valid_dates.append(current)
            current += timedelta(days=1)

        split_weeks = chunk_list_by_weeks(valid_dates)
        week_map[phase] = [[d.isoformat() for d in week] for week in split_weeks]

        for idx, week_dates in enumerate(split_weeks):
            if not week_dates:
                continue
            phase_week_to_continuous[(phase, idx)] = continuous_week_counter
            week_blocks.append({
                'phase': phase,
                'week_index': idx,
                'continuous_index': continuous_week_counter,
                'start_date': week_dates[0],
                'end_date': week_dates[-1]
            })
            continuous_week_counter += 1

    if request.method == "POST":
        action = request.POST.get('action', 'download')

        timetable_id = request.POST.get('timetable')
        if not timetable_id:
            messages.error(request, "Please select a timetable.")
            return redirect(request.path)

        try:
            timetable = Timetable.objects.get(id=timetable_id, department=user_department)
        except Timetable.DoesNotExist:
            messages.error(request, "Invalid timetable selected.")
            return redirect(request.path)

        selected_phases = request.POST.getlist('phases')
        selected_week_indices = request.POST.getlist('weeks')
        if not selected_phases or not selected_week_indices:
            messages.error(request, "Please select both phases and weeks.")
            return redirect(request.path)

        selected_week_nums = set()
        for phase in selected_phases:
            if phase == 'All':
                selected_week_nums = set(range(1, continuous_week_counter))
                break
            phase_weeks = week_map.get(phase, [])
            for idx_str in selected_week_indices:
                try:
                    idx = int(idx_str)
                    if 0 <= idx < len(phase_weeks):
                        selected_week_nums.add(phase_week_to_continuous[(phase, idx)])
                except Exception:
                    pass
        selected_week_nums = sorted(selected_week_nums)

        week_number_map = {wk: i+1 for i, wk in enumerate(selected_week_nums)}

        selected_dates_set = set()
        for wk_detail in week_blocks:
            if wk_detail['continuous_index'] in selected_week_nums:
                current = wk_detail['start_date']
                while current <= wk_detail['end_date']:
                    selected_dates_set.add(current)
                    current += timedelta(days=1)
        selected_dates = sorted(selected_dates_set)

        batches = Batch.objects.filter(department=user_department)

        wb = Workbook()
        wb.remove(wb.active)

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        color_fills = [
            PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid'),
            PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
            PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid'),
            PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'),
        ]

        # Helper to normalize lecture times/numbers for attendance matching.
        def _map_lec(val):
            time_map = {
                '8:45–9:45': '1', '8:45-9:45': '1',
                '9:45–10:45': '2', '9:45-10:45': '2',
                '11:30–12:30': '3', '11:30-12:30': '3',
                '12:30–13:30': '4', '12:30-13:30': '4',
                '13:45–14:45': '5', '13:45-14:45': '5'
            }
            v = str(val).replace('–', '-').replace('—', '-').replace(' ', '')
            for k in time_map:
                _k = k.replace('–', '-').replace('—', '-').replace(' ', '')
                if v == _k:
                    return time_map[k]
            try:
                return str(int(val))
            except Exception:
                return str(val).strip()

        

# ... inside your view, after creating workbook and fetching batches ...

        red_font = Font(color="FF0000")  # Red color font for Absent "A"

        for batch in batches:
            ws = wb.create_sheet(title=batch.name[:31])

            students = list(Student.objects.filter(department=user_department, batch=batch).order_by('roll_no'))
            n_students = len(students)

            # Write header for roll no and student name
            ws.cell(row=1, column=1, value='Roll No').font = Font(bold=True)
            ws.cell(row=1, column=2, value='Student Name').font = Font(bold=True)

            for idx, student in enumerate(students):
                ws.cell(row=3 + idx, column=1, value=student.roll_no)
                ws.cell(row=3 + idx, column=2, value=student.name)

            date_lecture_map = []
            col = 3
            week_cols_map = {}

            for dt in selected_dates:
                weekday = dt.strftime('%A').lower()
                lectures = list(TimetableEntry.objects.filter(
                    department=user_department, batch=batch,
                    day__iexact=weekday, timetable=timetable
                ).order_by('time'))

                continuous_idx = next((w['continuous_index'] for w in week_blocks if w['start_date'] <= dt <= w['end_date']), None)
                if continuous_idx is None:
                    continue
                color = color_fills[(continuous_idx - 1) % len(color_fills)]

                if lectures:
                    ws.merge_cells(
                        start_row=1,
                        start_column=col,
                        end_row=1,
                        end_column=col + len(lectures) - 1
                    )
                    date_cell = ws.cell(row=1, column=col)
                    date_cell.value = dt.strftime('%d-%b')
                    date_cell.font = Font(bold=True, color='FFFFFF')
                    date_cell.alignment = Alignment(horizontal='center', vertical='center')
                    date_cell.fill = color

                    for i, lec in enumerate(lectures, start=1):
                        lec_cell = ws.cell(row=2, column=col + i - 1)
                        lec_cell.value = f"Lect {i}\n{lec.subject.subject_name if lec.subject else 'N/A'}"
                        lec_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        lec_cell.fill = color
                        ws.column_dimensions[get_column_letter(col + i - 1)].width = 4
                        week_cols_map.setdefault(continuous_idx, []).append(col + i - 1)

                    date_lecture_map.append((dt, lectures))
                    col += len(lectures)
                else:
                    ws.cell(row=1, column=col, value=dt.strftime('%d-%b')).font = Font(bold=True, color='FFFFFF')
                    ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=1, column=col).fill = color
                    ws.cell(row=2, column=col).value = ""
                    ws.cell(row=2, column=col).fill = color
                    ws.column_dimensions[get_column_letter(col)].width = 4
                    week_cols_map.setdefault(continuous_idx, []).append(col)
                    date_lecture_map.append((dt, []))
                    col += 1

            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20

            ws.freeze_panes = 'C3'

            attendance_summary = defaultdict(lambda: defaultdict(lambda: {'held': 0, 'attended': 0}))

            for dt, lectures in date_lecture_map:
                wk_idx = next((w['continuous_index'] for w in week_blocks if w['start_date'] <= dt <= w['end_date']), None)
                if wk_idx not in selected_week_nums:
                    continue
                col_pos = 3
                for i, lec in enumerate(lectures, start=1):
                    # Match attendance by lecture_number = str(i), as in your summary logic
                    records = FacultyAttendance.objects.filter(batch=batch, date=dt, lecture_number=str(i))

                    absent_rolls = set()
                    for record in records:
                        if record.absent_roll_numbers:
                            absent_rolls.update(x.strip() for x in record.absent_roll_numbers.split(",") if x.strip())

                    for idx, student in enumerate(students):
                        is_absent = str(student.roll_no) in absent_rolls
                        cell = ws.cell(row=3 + idx, column=col_pos, value='A' if is_absent else 'P')
                        if is_absent:
                            cell.font = red_font
                    col_pos += 1

    # Rest of your summary generation and Excel setup (borders, widths, summary tables, cumulative, overall) remains untouched and exactly as your existing code...

    # Your final steps saving workbook to response or file, handling 'download' or 'whatsapp' actions, etc., remain unchanged.


            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20

            ws.freeze_panes = 'C3'

            attendance_summary = defaultdict(lambda: defaultdict(lambda: {'held': 0, 'attended': 0}))

            for dt, lectures in date_lecture_map:
                wk_idx = next((w['continuous_index'] for w in week_blocks if w['start_date'] <= dt <= w['end_date']), None)
                if wk_idx not in selected_week_nums:
                    continue
                col_pos = 3
                for lec in lectures:
                    # Map lecture time for attendance matching
                    lec_number = _map_lec(lec.time)
                    records = FacultyAttendance.objects.filter(batch=batch, date=dt)

                    absent_rolls = set()
                    for record in records:
                        if _map_lec(record.lecture_number) == lec_number:
                            if record.absent_roll_numbers:
                                absent_rolls.update(x.strip() for x in record.absent_roll_numbers.split(",") if x.strip())

                    subj_name = lec.subject.subject_name if lec.subject else "N/A"
                    for i, student in enumerate(students):
                        is_absent = str(student.roll_no) in absent_rolls
                        ws.cell(row=3 + i, column=col_pos, value='A' if is_absent else 'P')

                        s = attendance_summary[(wk_idx, subj_name)][student.roll_no]
                        s['held'] += 1
                        if not is_absent:
                            s['attended'] += 1
                    col_pos += 1

            max_row = 3 + len(students) - 1
            max_col = col - 1
            for r in range(1, max_row + 1):
                for c_ in range(1, max_col + 1):
                    ws.cell(r, c_).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                   top=Side(style='thin'), bottom=Side(style='thin'))

            summary_start = max_row + 3
            ws.cell(summary_start, 1, 'Roll No').font = Font(bold=True)
            ws.cell(summary_start, 2, 'Student Name').font = Font(bold=True)
            ws.cell(summary_start - 1, 1, 'Summary').font = Font(bold=True)
            ws.merge_cells(start_row=summary_start - 1, start_column=1, end_row=summary_start - 1, end_column=2)
            ws.cell(summary_start - 1, 1).alignment = Alignment(horizontal='center')

            subjects = sorted(set(sub for _, sub in attendance_summary.keys()))
            weeks_sorted = selected_week_nums
            week_display_map = {w: idx + 1 for idx, w in enumerate(weeks_sorted)}

            col_pos = 3
            for idx, wk in enumerate(weeks_sorted):
                fill = color_fills[idx % len(color_fills)]
                display_wk = week_display_map[wk]
                for subj in subjects:
                    ws.cell(summary_start - 1, col_pos, f"Week {display_wk} - {subj}").font = Font(bold=True)
                    ws.cell(summary_start - 1, col_pos).alignment = Alignment(horizontal='center')
                    ws.cell(summary_start - 1, col_pos).fill = fill
                    ws.merge_cells(start_row=summary_start - 1, start_column=col_pos, end_row=summary_start - 1,
                                   end_column=col_pos + 2)
                    col_pos += 3

            col_pos = 3
            subhdr_row = summary_start
            for idx, wk in enumerate(weeks_sorted):
                fill = color_fills[idx % len(color_fills)]
                for _ in subjects:
                    for i, label in enumerate(['Held', 'Attended', 'Percentage']):
                        cell = ws.cell(subhdr_row, col_pos + i, label)
                        cell.font = Font(bold=True)
                        cell.fill = fill
                    col_pos += 3

            data_start = subhdr_row + 1
            for idx, student in enumerate(students):
                ws.cell(data_start + idx, 1, student.roll_no)
                ws.cell(data_start + idx, 2, student.name)
                col_pos = 3
                for wk in weeks_sorted:
                    for subj in subjects:
                        data = attendance_summary.get((wk, subj), {}).get(student.roll_no, {'held': 0, 'attended': 0})
                        held, attended = data['held'], data['attended']
                        perc = (attended / held * 100) if held else 0
                        ws.cell(data_start + idx, col_pos, held)
                        ws.cell(data_start + idx, col_pos + 1, attended)
                        perc_cell = ws.cell(data_start + idx, col_pos + 2, f'{perc:.1f}%')
                        perc_cell.alignment = Alignment(horizontal='center')
                        col_pos += 3

            cum_col_start = col_pos
            ws.merge_cells(start_row=subhdr_row - 1, start_column=cum_col_start, end_row=subhdr_row - 1,
                           end_column=cum_col_start + 3 * len(subjects) - 1)
            cum_header = ws.cell(subhdr_row - 1, cum_col_start, 'Cumulative Summary')
            cum_header.font = Font(bold=True)
            cum_header.alignment = Alignment(horizontal='center')

            ws.cell(subhdr_row, cum_col_start, 'Held').font = Font(bold=True)
            ws.cell(subhdr_row, cum_col_start + 1, 'Attended').font = Font(bold=True)
            ws.cell(subhdr_row, cum_col_start + 2, 'Percentage').font = Font(bold=True)

            cum_data_start = subhdr_row + 1
            for idx, student in enumerate(students):
                ws.cell(cum_data_start + idx, 1, student.roll_no)
                ws.cell(cum_data_start + idx, 2, student.name)
                col_ptr = cum_col_start
                for subj in subjects:
                    total_held = sum(attendance_summary.get((wk, subj), {}).get(student.roll_no, {'held': 0})['held'] for wk in weeks_sorted)
                    total_attended = sum(attendance_summary.get((wk, subj), {}).get(student.roll_no, {'attended': 0})['attended'] for wk in weeks_sorted)
                    total_perc = (total_attended / total_held * 100) if total_held else 0
                    ws.cell(cum_data_start + idx, col_ptr, total_held)
                    ws.cell(cum_data_start + idx, col_ptr + 1, total_attended)
                    perc_cell = ws.cell(cum_data_start + idx, col_ptr + 2, f'{total_perc:.1f}%')
                    perc_cell.alignment = Alignment(horizontal='center')
                    if total_perc < 75:
                        perc_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    col_ptr += 3

            overall_col_start = col_ptr
            ws.merge_cells(start_row=subhdr_row - 1, start_column=overall_col_start, end_row=subhdr_row - 1,
                           end_column=overall_col_start + 2)
            overall_hdr_cell = ws.cell(subhdr_row - 1, overall_col_start, 'Overall Total Summary')
            overall_hdr_cell.font = Font(bold=True)
            overall_hdr_cell.alignment = Alignment(horizontal='center')

            ws.cell(subhdr_row, overall_col_start, 'Held').font = Font(bold=True)
            ws.cell(subhdr_row, overall_col_start + 1, 'Attended').font = Font(bold=True)
            ws.cell(subhdr_row, overall_col_start + 2, 'Percentage').font = Font(bold=True)

            overall_data_start = subhdr_row + 1
            for idx, student in enumerate(students):
                ws.cell(overall_data_start + idx, 1, student.roll_no)
                ws.cell(overall_data_start + idx, 2, student.name)
                total_held = 0
                total_attended = 0
                for subj in subjects:
                    total_held += sum(attendance_summary.get((wk, subj), {}).get(student.roll_no, {'held': 0})['held'] for wk in weeks_sorted)
                    total_attended += sum(attendance_summary.get((wk, subj), {}).get(student.roll_no, {'attended': 0})['attended'] for wk in weeks_sorted)
                total_perc = (total_attended / total_held * 100) if total_held else 0
                ws.cell(overall_data_start + idx, overall_col_start, total_held)
                ws.cell(overall_data_start + idx, overall_col_start + 1, total_attended)
                perc_cell = ws.cell(overall_data_start + idx, overall_col_start + 2, f'{total_perc:.1f}%')
                perc_cell.alignment = Alignment(horizontal='center')
                if total_perc < 75:
                    perc_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            max_col_idx = overall_col_start + 2
            for cidx in range(1, max_col_idx + 1):
                if cidx == 1:
                    ws.column_dimensions[get_column_letter(cidx)].width = 12
                elif cidx == 2:
                    ws.column_dimensions[get_column_letter(cidx)].width = 20
                else:
                    ws.column_dimensions[get_column_letter(cidx)].width = 6

            last_summary_row = overall_data_start + n_students - 1
            for r in range(subhdr_row - 4, last_summary_row + 1):
                for c in range(1, max_col_idx + 1):
                    ws.cell(r, c).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                  top=Side(style='thin'), bottom=Side(style='thin'))

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'Attendance_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        if action == 'download':
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        elif action == 'whatsapp':
            phone_number = request.POST.get('phone_number', '').strip()

            if not phone_number:
                messages.error(request, "Phone number is required to send on WhatsApp.")
                return redirect(request.path)

            save_dir = os.path.join(settings.MEDIA_ROOT, 'attendance_reports')
            os.makedirs(save_dir, exist_ok=True)
            file_path = os.path.join(save_dir, filename)

            with open(file_path, 'wb') as f:
                f.write(output.read())

            public_file_url = request.build_absolute_uri(settings.MEDIA_URL + 'attendance_reports/' + filename)
            encoded_url = quote(public_file_url)

            whatsapp_url = f"https://wa.me/{phone_number}?text=Please%20find%20the%20attendance%20report%20here:%20{encoded_url}"

            return redirect(whatsapp_url)

    # For GET or other methods, render normally
    return render(request, 'attendance/attendance_sheet_manager.html', {
        'timetables': timetables,
        'phases': phases,
        'week_map': week_map,
        'selected_timetable': None,
        'session_window': session_window,
    })

from django.shortcuts import render

