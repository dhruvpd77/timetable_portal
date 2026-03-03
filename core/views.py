from collections import defaultdict
from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from core.utils.timetable_solver import generate_timetable, diagnose_timetable_failure
from .models import (
    College, Department, Batch, Lab, Room, CourseAssignment, CourseSpec,
    Faculty, Timetable, TimetableEntry, VisitingFacultyBlock,
    TimeSettings, GlobalDay, GlobalTimeSlot, BatchRoomLabMapping, TimetableType,
    FacultyPreferredSlot, LeaveRequest, LeaveReassignment,
    FacultyBlock, RoomBlock, LabBlock
)
from .forms import FacultyPreferredSlotForm
from .forms import BatchForm, LabForm, RoomForm, CourseAssignmentForm, CourseSpecForm, TimetableTypeForm, AdminPasswordForm, CollegeForm, DepartmentForm, UserForm, UploadExcelForm, BatchRoomLabSimpleMappingForm, UploadExcelTimetableForm



def welcome_view(request):
    colleges = College.objects.all()
    departments = Department.objects.none()

    selected_college_id = request.POST.get('college') or request.GET.get('college')

    if selected_college_id:
        departments = Department.objects.filter(college_id=selected_college_id)

    if request.method == 'POST' and 'login_button' in request.POST:
        college_id = request.POST.get('college')
        dept_id = request.POST.get('department')
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user:
            try:
                department = Department.objects.get(id=dept_id, user=user, college_id=college_id)
                login(request, user)
                # Store in session
                request.session['selected_college_id'] = college_id
                request.session['selected_college_name'] = department.college.name
                request.session['selected_department_id'] = dept_id
                request.session['selected_department_name'] = department.name
                messages.success(request, "Login successful.")
                return redirect('department_time_settings')  # or your dashboard view
            except Department.DoesNotExist:
                messages.error(request, "Invalid department or college selection.")
        else:
            messages.error(request, "Invalid username or password.")

    return render(request, 'core/welcome.html', {
        'colleges': colleges,
        'departments': departments,
        'selected_college_id': selected_college_id
    })



@login_required
def logout_view(request):

    logout(request)
    return redirect("welcome")

@login_required
def admin_setup_view(request):
    if request.method == "POST":
        if "add_college" in request.POST:
            name = request.POST.get("college_name")
            if name:
                College.objects.get_or_create(name=name)
                messages.success(request, f"College '{name}' added.")

        elif "add_department" in request.POST:
            dept_name = request.POST.get("dept_name")
            college_id = request.POST.get("dept_college")
            if dept_name and college_id:
                college = College.objects.get(id=college_id)
                Department.objects.get_or_create(name=dept_name, college=college)
                messages.success(request, f"Department '{dept_name}' added to {college.name}.")

    colleges = College.objects.all()
    departments = Department.objects.select_related('college').all()

    return render(request, "core/admin_setup.html", {
        "colleges": colleges,
        "departments": departments,
    })
import json

@login_required
def manage_visiting_blocks(request):
    user = request.user
    user_dept = Department.objects.filter(user=user).first()
    if not user_dept:
        messages.error(request, "You are not linked to any department.")
        return redirect("home")

    faculties = Faculty.objects.filter(department=user_dept).order_by("short_name")
    selected_faculty_name = request.GET.get("faculty")
    selected_ext_dept_id = request.GET.get("engaged_department")
    selected_tt_id = request.GET.get("engaged_timetable")

    engaged_departments, engaged_timetables, blocked_slots = [], [], []
    show_slots = False

    # ✅ Fetch all blocks made by THIS department
    all_blocks = VisitingFacultyBlock.objects.filter(main_department=user_dept)

    if selected_faculty_name:
        fac_obj = Faculty.objects.filter(short_name=selected_faculty_name).first()

        engaged_departments = Department.objects.exclude(id=user_dept.id)
        if selected_ext_dept_id:
            engaged_timetables = Timetable.objects.filter(department_id=selected_ext_dept_id)
            if selected_tt_id:
                entries = TimetableEntry.objects.filter(
                    timetable_id=selected_tt_id,
                    faculty__short_name=selected_faculty_name
                )
                blocked_slots = [(e.day, e.time) for e in entries]
                show_slots = True

    # ✅ Handle Block POST
    if request.method == "POST" and "block_slots" in request.POST:
        ext_dept_id = request.POST.get("engaged_department_id")
        timetable_id = request.POST.get("engaged_timetable_id")
        selected_faculty_name = request.POST.get("selected_faculty_name")

        fac_obj = Faculty.objects.filter(short_name=selected_faculty_name).first()
        if not fac_obj:
            messages.error(request, "Faculty not found during block.")
            return redirect("manage_visiting_blocks")

        # Get blocked entries from selected timetable
        entries = TimetableEntry.objects.filter(timetable_id=timetable_id, faculty__short_name=selected_faculty_name)
        blocked = [(e.day, e.time) for e in entries]

        # ✅ Save block under logged-in department (user_dept)
        VisitingFacultyBlock.objects.filter(faculty=fac_obj, main_department=user_dept).delete()
        VisitingFacultyBlock.objects.create(
            faculty=fac_obj,
            main_department=user_dept,
            blocked_slots=json.dumps(blocked)
        )
        messages.success(request, "Blocked visiting faculty slots successfully.")
        return redirect("manage_visiting_blocks")

    # ✅ Handle Delete POST
    if request.method == "POST" and "delete_block_id" in request.POST:
        block_id = request.POST.get("delete_block_id")
        VisitingFacultyBlock.objects.filter(id=block_id, main_department=user_dept).delete()
        messages.success(request, "Deleted visiting faculty block.")
        return redirect("manage_visiting_blocks")

    context = {
        "faculties": faculties,
        "selected_faculty_name": selected_faculty_name,
        "selected_ext_dept_id": selected_ext_dept_id,
        "selected_tt_id": selected_tt_id,
        "engaged_departments": engaged_departments,
        "engaged_timetables": engaged_timetables,
        "blocked_slots": blocked_slots,
        "show_slots": show_slots,
        "all_blocks": all_blocks,
    }

    return render(request, "core/manage_visiting_blocks.html", context)

from .models import GlobalDay, GlobalTimeSlot
from datetime import time

@login_required
def global_time_settings_view(request):
    if request.method == "POST":
        if "add_day" in request.POST:
            day_name = request.POST.get("day_name")
            if day_name:
                GlobalDay.objects.get_or_create(name=day_name)
                messages.success(request, f"Day '{day_name}' added.")

        elif "add_slot" in request.POST:
            start = request.POST.get("start_time")
            end = request.POST.get("end_time")
            if start and end:
                start_time = time.fromisoformat(start)
                end_time = time.fromisoformat(end)
                GlobalTimeSlot.objects.get_or_create(start_time=start_time, end_time=end_time)
                messages.success(request, f"Slot {start}–{end} added.")

    # Delete actions
    if request.method == "POST" and "delete_day" in request.POST:
        GlobalDay.objects.filter(id=request.POST.get("delete_day")).delete()
        messages.success(request, "Day deleted.")

    if request.method == "POST" and "delete_slot" in request.POST:
        GlobalTimeSlot.objects.filter(id=request.POST.get("delete_slot")).delete()
        messages.success(request, "Time slot deleted.")

    days = GlobalDay.objects.all().order_by('id')
    slots = GlobalTimeSlot.objects.all().order_by('start_time')

    return render(request, "core/global_time_settings.html", {
        "days": days,
        "slots": slots
    })
from django.contrib.auth.decorators import login_required
from .models import Department, GlobalDay, GlobalTimeSlot, TimeSettings
from django.forms import ModelMultipleChoiceField

@login_required
def department_time_settings_view(request):
    user = request.user
    try:
        dept = Department.objects.get(user=user)
    except Department.DoesNotExist:
        messages.error(request, "No department found for this user.")
        return redirect('welcome')

    global_days = GlobalDay.objects.all()
    global_slots = GlobalTimeSlot.objects.all()

    timesettings, _ = TimeSettings.objects.get_or_create(department=dept)

    if request.method == "POST":
        selected_day_ids = request.POST.getlist("selected_days")
        selected_slot_ids = request.POST.getlist("selected_slots")
        break_slot_ids = request.POST.getlist("break_slots")

        timesettings.selected_days.set(selected_day_ids)
        timesettings.selected_slots.set(selected_slot_ids)
        timesettings.break_slots.set(break_slot_ids)

        messages.success(request, "Time settings saved.")
        return redirect('department_time_settings')

    return render(request, "core/department_time_settings.html", {
        "global_days": global_days,
        "global_slots": global_slots,
        "selected_days": timesettings.selected_days.all(),
        "selected_slots": timesettings.selected_slots.all(),
        "break_slots": timesettings.break_slots.all()
    })
from .models import Faculty

@login_required
def faculty_management_view(request):
    try:
        dept = Department.objects.get(user=request.user)
    except Department.DoesNotExist:
        messages.error(request, "No department linked to this login.")
        return redirect('welcome')

    if request.method == "POST":
        if "add_faculty" in request.POST:
            full_name = request.POST.get("full_name")
            short_name = request.POST.get("short_name")
            default_load = request.POST.get("default_load")
            if full_name and short_name:
                Faculty.objects.create(
                    department=dept,
                    full_name=full_name,
                    short_name=short_name,
                    default_load=default_load or 18
                )
                messages.success(request, "Faculty added successfully.")
                return redirect('faculty_management')

        if "delete_faculty" in request.POST:
            fid = request.POST.get("delete_faculty")
            Faculty.objects.filter(id=fid, department=dept).delete()
            messages.success(request, "Faculty deleted.")
            return redirect('faculty_management')

    faculty_list = Faculty.objects.filter(department=dept).order_by('short_name')

    return render(request, "core/faculty_management.html", {
        "faculty_list": faculty_list
    })
@login_required
def edit_faculty_view(request, faculty_id):
    try:
        dept = Department.objects.get(user=request.user)
    except Department.DoesNotExist:
        messages.error(request, "Department not found.")
        return redirect('welcome')

    faculty = get_object_or_404(Faculty, id=faculty_id, department=dept)

    if request.method == 'POST':
        faculty.full_name = request.POST.get('full_name')
        faculty.short_name = request.POST.get('short_name')
        faculty.default_load = request.POST.get('default_load') or 0
        faculty.save()
        messages.success(request, "Faculty updated successfully.")
        return redirect('faculty_management')

    return render(request, 'core/edit_faculty.html', {
        'faculty': faculty
    })
@login_required
def batch_management_view(request):
    try:
        dept = Department.objects.get(user=request.user)
    except Department.DoesNotExist:
        messages.error(request, "Department not found.")
        return redirect('welcome')

    batches = Batch.objects.filter(department=dept).order_by('name')

    if request.method == 'POST':
        if 'delete_batch' in request.POST:
            batch_id = request.POST.get('delete_batch')
            Batch.objects.filter(id=batch_id, department=dept).delete()
            messages.success(request, "Batch deleted.")
            return redirect('batch_management')

        form = BatchForm(request.POST)
        if form.is_valid():
            new_batch = form.save(commit=False)
            new_batch.department = dept
            new_batch.save()
            messages.success(request, "Batch added.")
            return redirect('batch_management')
    else:
        form = BatchForm()

    return render(request, 'core/batch_management.html', {
        'form': form,
        'batches': batches
    })

@login_required
def edit_batch_view(request, batch_id):
    try:
        dept = Department.objects.get(user=request.user)
        batch = Batch.objects.get(id=batch_id, department=dept)
    except (Department.DoesNotExist, Batch.DoesNotExist):
        messages.error(request, "Batch not found.")
        return redirect('batch_management')

    if request.method == 'POST':
        form = BatchForm(request.POST, instance=batch)
        if form.is_valid():
            form.save()
            messages.success(request, "Batch updated.")
            return redirect('batch_management')
    else:
        form = BatchForm(instance=batch)

    return render(request, 'core/edit_batch.html', {
        'form': form,
        'batch': batch
    })
@login_required
def lab_management_view(request):
    try:
        dept = Department.objects.get(user=request.user)
    except Department.DoesNotExist:
        messages.error(request, "Department not found.")
        return redirect('welcome')

    labs = Lab.objects.filter(department=dept).order_by('name')

    if request.method == 'POST':
        if 'delete_lab' in request.POST:
            lab_id = request.POST.get('delete_lab')
            Lab.objects.filter(id=lab_id, department=dept).delete()
            messages.success(request, "Lab deleted.")
            return redirect('lab_management')

        # ✅ This part was missing
        form = LabForm(request.POST)
        if form.is_valid():
            new_lab = form.save(commit=False)
            new_lab.department = dept     # ✅ Set department!
            new_lab.save()
            messages.success(request, "Lab added successfully.")
            return redirect('lab_management')
    else:
        form = LabForm()

    return render(request, 'core/lab_management.html', {
        'form': form,
        'labs': labs
    })


@login_required
def edit_lab_view(request, lab_id):
    try:
        dept = Department.objects.get(user=request.user)
        lab = Lab.objects.get(id=lab_id, department=dept)
    except (Department.DoesNotExist, Lab.DoesNotExist):
        messages.error(request, "Lab not found.")
        return redirect('lab_management')

    if request.method == 'POST':
        form = LabForm(request.POST, instance=lab)
        if form.is_valid():
            form.save()
            messages.success(request, "Lab updated.")
            return redirect('lab_management')
    else:
        form = LabForm(instance=lab)

    return render(request, 'core/edit_lab.html', {
        'form': form,
        'lab': lab
    })
@login_required
def room_management_view(request):
    try:
        dept = Department.objects.get(user=request.user)
    except Department.DoesNotExist:
        messages.error(request, "Department not found.")
        return redirect('welcome')

    rooms = Room.objects.filter(department=dept).order_by('name')

    if request.method == 'POST':
        if 'delete_room' in request.POST:
            room_id = request.POST.get('delete_room')
            Room.objects.filter(id=room_id, department=dept).delete()
            messages.success(request, "Room deleted.")
            return redirect('room_management')

        form = RoomForm(request.POST)
        if form.is_valid():
            new_room = form.save(commit=False)
            new_room.department = dept
            new_room.save()
            messages.success(request, "Room added.")
            return redirect('room_management')
    else:
        form = RoomForm()

    return render(request, 'core/room_management.html', {
        'form': form,
        'rooms': rooms
    })


@login_required
def edit_room_view(request, room_id):
    try:
        dept = Department.objects.get(user=request.user)
        room = Room.objects.get(id=room_id, department=dept)
    except (Department.DoesNotExist, Room.DoesNotExist):
        messages.error(request, "Room not found.")
        return redirect('room_management')

    if request.method == 'POST':
        form = RoomForm(request.POST, instance=room)
        if form.is_valid():
            form.save()
            messages.success(request, "Room updated.")
            return redirect('room_management')
    else:
        form = RoomForm(instance=room)

    return render(request, 'core/edit_room.html', {
        'form': form,
        'room': room
    })
# core/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Department, CourseSpec, CourseAssignment, Faculty, Batch
# from .forms import CourseSpecForm, CourseAssignmentForm

# ------------------------------------------
# COURSE SPECIFICATION VIEWS
# ------------------------------------------

@login_required
def course_specification_view(request):
    department = get_object_or_404(Department, user=request.user)
    form = CourseSpecForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            spec = form.save(commit=False)
            spec.department = department
            spec.save()
            messages.success(request, "Course Specification Added!")
            return redirect('course_specification')

    specs = CourseSpec.objects.filter(department=department)
    return render(request, 'core/course_specification.html', {'form': form, 'specs': specs})


@login_required
def edit_course_specification(request, spec_id):
    department = get_object_or_404(Department, user=request.user)
    spec = get_object_or_404(CourseSpec, id=spec_id, department=department)
    form = CourseSpecForm(request.POST or None, instance=spec)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Course Specification Updated.")
            return redirect('course_specification')

    return render(request, 'core/edit_course_specification.html', {'form': form, 'spec': spec})


@login_required
def delete_course_specification(request, spec_id):
    department = get_object_or_404(Department, user=request.user)
    spec = get_object_or_404(CourseSpec, id=spec_id, department=department)
    spec.delete()
    messages.success(request, "Course Specification Deleted.")
    return redirect('course_specification')

# ------------------------------------------
# COURSE ASSIGNMENT VIEWS (with MULTIPLE BATCHES)
# ------------------------------------------

# core/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import CourseAssignment, CourseSpec, Faculty, Batch, Department
# from .forms import CourseAssignmentForm
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError

from django.shortcuts import render, redirect, get_object_or_404
from .models import CourseAssignment, CourseSpec, Faculty, Batch
# from .forms import CourseAssignmentForm
from django.contrib import messages
from django.db import IntegrityError

# core/views.py
from django.shortcuts import render, redirect, get_object_or_404
from .models import CourseAssignment, CourseSpec, Faculty, Batch, Department
# from .forms import CourseAssignmentForm
from django.contrib import messages

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import CourseAssignment, CourseSpec
# from .forms import CourseAssignmentForm
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import CourseAssignment, Department
# from .forms import CourseAssignmentForm

@login_required
def assign_course_view(request):
    department = get_object_or_404(Department, user=request.user)

    if request.method == "POST":
        form = CourseAssignmentForm(request.POST, department=department)
        if form.is_valid():
            subject = form.cleaned_data['subject']
            faculty = form.cleaned_data['faculty']
            batch = form.cleaned_data['batch']
            hours = form.cleaned_data['hours']
            room_or_lab = form.cleaned_data['room_or_lab']

            CourseAssignment.objects.create(
                department=department,
                subject=subject,
                faculty=faculty,
                batch=batch,
                hours=hours,
                room_or_lab=room_or_lab
            )
            messages.success(request, "Course assignment saved.")
            return redirect("assign_course")
    else:
        form = CourseAssignmentForm(department=department)

    assignments = CourseAssignment.objects.filter(department=department)
    return render(request, "core/assign_course.html", {
        "form": form,
        "assignments": assignments
    })

@login_required
def delete_assignment_view(request, pk):
    assignment = get_object_or_404(CourseAssignment, pk=pk)
    assignment.delete()
    messages.success(request, "Assignment deleted.")
    return redirect("assign_course")

@login_required
def edit_assignment_view(request, pk):
    assignment = get_object_or_404(CourseAssignment, pk=pk)
    department = get_object_or_404(Department, user=request.user)

    if request.method == "POST":
        form = CourseAssignmentForm(request.POST, department=department)
        if form.is_valid():
            subject = form.cleaned_data['subject']
            faculty = form.cleaned_data['faculty']
            batch = form.cleaned_data['batch']
            hours = form.cleaned_data['hours']
            room_or_lab = form.cleaned_data['room_or_lab']

            # Update the existing assignment
            assignment.subject = subject
            assignment.faculty = faculty
            assignment.batch = batch
            assignment.hours = hours
            assignment.room_or_lab = room_or_lab
            assignment.save()

            messages.success(request, "Assignment updated.")
            return redirect("assign_course")
    else:
        form = CourseAssignmentForm(department=department, initial={
            'subject': assignment.subject,
            'faculty': assignment.faculty,
            'batch': assignment.batch,
            'hours': assignment.hours,
            'room_or_lab': assignment.room_or_lab,
        })

    return render(request, "core/edit_assignment.html", {
        "form": form,
        "assignment": assignment
    })
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import (
    Department, CourseSpec, CourseAssignment, Faculty, Batch, Room, Lab, TimeSettings,
    Timetable, TimetableEntry
)
from .utils.timetable_solver import generate_timetable




from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import (
    Department, CourseSpec, CourseAssignment, Faculty, Batch, Room, Lab, TimeSettings,
    Timetable, TimetableEntry, VisitingFacultyBlock
)
from .utils.timetable_solver import generate_timetable
import json

def get_blocked_faculty_slots(department):
    """
    Returns a dictionary mapping faculty short_name to a set of (day, slot)
    for all slots blocked for that faculty (for current department).
    """
    blocks = VisitingFacultyBlock.objects.filter(faculty__department=department)
    blocked = {}
    for block in blocks:
        faculty_name = block.faculty.short_name
        try:
            slot_list = json.loads(block.blocked_slots)
            blocked.setdefault(faculty_name, set()).update(tuple(pair) for pair in slot_list)
        except Exception:
            continue
    return blocked


def get_blocked_room_slots(department):
    """
    Returns a dictionary mapping room name to a set of (day, slot)
    for all slots blocked for that room (for current department).
    """
    blocks = RoomBlock.objects.filter(department=department)
    blocked = {}
    for block in blocks:
        room_name = block.room.name
        try:
            slot_list = block.blocked_slots
            blocked.setdefault(room_name, set()).update(tuple(pair) for pair in slot_list)
        except Exception:
            continue
    return blocked


def get_blocked_lab_slots(department):
    """
    Returns a dictionary mapping lab name to a set of (day, slot)
    for all slots blocked for that lab (for current department).
    """
    blocks = LabBlock.objects.filter(department=department)
    blocked = {}
    for block in blocks:
        lab_name = block.lab.name
        try:
            slot_list = block.blocked_slots
            blocked.setdefault(lab_name, set()).update(tuple(pair) for pair in slot_list)
        except Exception:
            continue
    return blocked


from collections import defaultdict
from django.contrib import messages

@login_required
def generate_timetable_view(request):
    department = get_object_or_404(Department, user=request.user)
    settings = get_object_or_404(TimeSettings, department=department)
    specs = CourseSpec.objects.filter(department=department)
    assignments = CourseAssignment.objects.filter(department=department)
    faculty_objs = {f.id: f for f in Faculty.objects.filter(department=department)}
    batch_objs = {b.id: b for b in Batch.objects.filter(department=department)}
    rooms = [r.name for r in Room.objects.filter(department=department)]
    labs = [l.name for l in Lab.objects.filter(department=department)]



    # Fetch batch->room and batch->lab mappings (dict[str, set[str]])
    mappings_qs = BatchRoomLabMapping.objects.filter(department=department).select_related('batch', 'room', 'lab')

    batch_to_rooms = defaultdict(set)
    batch_to_labs = defaultdict(set)
    for mapping in mappings_qs:
        batch_name = mapping.batch.name
        if mapping.room:
            batch_to_rooms[batch_name].add(mapping.room.name)
        if mapping.lab:
            batch_to_labs[batch_name].add(mapping.lab.name)

    timetable_by_day_slot = None
    result_entries = []
    gen_error = None
    gen_reasons = []
    try:
        # Fetch blocked slots if you use this feature
        blocked_faculty_slots = get_blocked_faculty_slots(department)
        blocked_room_slots = get_blocked_room_slots(department)
        blocked_lab_slots = get_blocked_lab_slots(department)

        # Merge FacultyBlock (direct blocks) into blocked_faculty_slots
        for fb in FacultyBlock.objects.filter(department=department):
            fac = fb.faculty.short_name
            for item in (fb.blocked_slots or []):
                if isinstance(item, dict):
                    day, slot = item.get("day"), item.get("slot")
                elif isinstance(item, (list, tuple)) and len(item) >= 2:
                    day, slot = item[0], item[1]
                else:
                    day, slot = None, None
                if day and slot:
                    blocked_faculty_slots.setdefault(fac, set()).add((day, slot))

        # Fetch preferred slots
        preferred_faculty_slots = get_preferred_faculty_slots(department)

        # Fetch timetable type setting
        try:
            timetable_type_obj = TimetableType.objects.get(department=department)
            timetable_type = timetable_type_obj.slot_type
        except TimetableType.DoesNotExist:
            timetable_type = '2_hour'  # Default to 2-hour pair slots

        timetable_by_day_slot, result_entries, status, status_name = generate_timetable(
            department, settings, specs, assignments, faculty_objs, batch_objs,
            rooms, labs,
            batch_to_rooms=batch_to_rooms,
            batch_to_labs=batch_to_labs,
            blocked_faculty_slots=blocked_faculty_slots,
            blocked_room_slots=blocked_room_slots,
            blocked_lab_slots=blocked_lab_slots,
            preferred_faculty_slots=preferred_faculty_slots,
            timetable_type=timetable_type
        )

        # If generation failed (INFEASIBLE), run diagnostic to find likely causes
        from django.urls import reverse
        if status_name not in ("OPTIMAL", "FEASIBLE") and timetable_by_day_slot is None:
            gen_error = f"Timetable could not be generated (solver status: {status_name})."
            gen_reasons = diagnose_timetable_failure(
                department, settings, specs, assignments, faculty_objs, batch_objs,
                rooms, labs,
                batch_to_rooms=batch_to_rooms,
                batch_to_labs=batch_to_labs,
                blocked_faculty_slots=blocked_faculty_slots,
                blocked_room_slots=blocked_room_slots,
                blocked_lab_slots=blocked_lab_slots,
                preferred_faculty_slots=preferred_faculty_slots,
                timetable_type=timetable_type,
            )
            # Add URLs for quick navigation
            for r in gen_reasons:
                if r.get("url_name"):
                    r["url"] = reverse(r["url_name"])
                if r.get("url_name2"):
                    r["url2"] = reverse(r["url_name2"])
    except Exception as e:
        import traceback
        gen_error = f"Timetable generation failed: {e}"
        traceback.print_exc()

    if request.method == "POST" and "generate" in request.POST:
        # You can optionally re-run generate_timetable on demand here or above
        request.session["timetable_result_entries"] = result_entries
    elif request.method == "POST" and "save" in request.POST:
        result_entries = request.session.get("timetable_result_entries", [])
        name = request.POST.get("timetable_name") or "Default Timetable"
        timetable = Timetable.objects.create(
            department=department, name=name, created_by=request.user
        )
        for entry in result_entries:
            subject = CourseSpec.objects.get(subject_name=entry["subject"], department=department)
            faculty = Faculty.objects.get(short_name=entry["faculty"], department=department)
            batch = Batch.objects.get(name=entry["batch"], department=department)
            room = Room.objects.get(name=entry["room"], department=department) if entry["room"] else None
            lab = Lab.objects.get(name=entry["lab"], department=department) if entry["lab"] else None
            TimetableEntry.objects.create(
                timetable=timetable,
                department=department,
                subject=subject,
                faculty=faculty,
                batch=batch,
                day=entry["day"],
                time=entry["time"],
                room=room,
                lab=lab
            )
        messages.success(request, "Timetable saved successfully!")
        return redirect("generate_timetable")

    return render(request, "core/generate_timetable.html", {
        "timetable_by_day_slot": timetable_by_day_slot,
        "slots": [f"{s.start_time.strftime('%H:%M')}–{s.end_time.strftime('%H:%M')}" for s in settings.selected_slots.all()],
        "show_save": bool(result_entries),
        "gen_error": gen_error,
        "gen_reasons": gen_reasons,
    })


from django.shortcuts import render, get_object_or_404
from .models import Timetable, TimetableEntry, Department, TimeSettings, Batch
import json

from django.shortcuts import render
from .models import Timetable, TimetableEntry, TimeSettings, Batch

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
import re

@login_required
def view_past_timetable(request):
    department = request.user.department  
    timetables = Timetable.objects.filter(department=department).order_by('-created_at')
    tt_id = request.GET.get('tt_id')
    selected_tt = None
    entries = []
    grid = {}
    days_order = []
    slots_order = []
    batches = []

    if tt_id:
        selected_tt = timetables.get(id=tt_id)
        entries = TimetableEntry.objects.filter(
            timetable=selected_tt
        ).select_related('batch', 'subject', 'faculty', 'room', 'lab')

        # Get TimeSettings
        ts = TimeSettings.objects.get(department=department)
        days_order = [d.name for d in ts.selected_days.all()]
        slots_order = [
            f"{slot.start_time.strftime('%H:%M')} – {slot.end_time.strftime('%H:%M')}"
            for slot in ts.selected_slots.all()
        ]
        batches = [b.name for b in Batch.objects.filter(department=department)]

        # Normalize day (Mon → Monday)
        def normalize_day(day_val):
            if not day_val:
                return day_val
            day_abbr = day_val.strip()[:3].lower()
            for d in days_order:
                if d.lower()[:3] == day_abbr:
                    return d
            return day_val

        # Normalize time slot (8:45 to 9:45 → 08:45 – 09:45)
        def normalize_time(time_val):
            if not time_val:
                return time_val
            s = str(time_val)
            # Split by any dash/to variation
            parts = re.split(r"to|-|–|—", s)
            if len(parts) != 2:
                return s.strip()
            def pad(t): return ":".join([f"{int(x):02d}" for x in t.strip().split(":")])
            start, end = pad(parts[0]), pad(parts[1])
            candidate = f"{start} – {end}"
            for slot in slots_order:
                if candidate.replace(" ", "") == slot.replace(" ", ""):
                    return slot
            return s.strip()

        # Create empty grid
        grid = {
            day: {slot: {batch: "" for batch in batches} for slot in slots_order}
            for day in days_order
        }

        # Fill grid with normalized values
        for entry in entries:
            day = normalize_day(entry.day)
            time = normalize_time(entry.time)
            batch = entry.batch.name
            if day in grid and time in grid[day] and batch in grid[day][time]:
                subj = entry.subject.subject_name
                fac = entry.faculty.full_name
                room_or_lab_name = entry.room.name if entry.room else (entry.lab.name if entry.lab else "")
                cell_text = f"{subj} ({fac})"
                if room_or_lab_name:
                    cell_text += f" ({room_or_lab_name})"
                if grid[day][time][batch]:
                    grid[day][time][batch] += "<br>" + cell_text
                else:
                    grid[day][time][batch] = cell_text

    

    if request.GET.get("download") == "facultywise" and selected_tt:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from collections import defaultdict

        wb = Workbook()
        ws = wb.active
        ws.title = "Facultywise Timetable"

        # Styling
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style="medium", color="000000"),
            right=Side(style="medium", color="000000"),
            top=Side(style="medium", color="000000"),
            bottom=Side(style="medium", color="000000")
        )
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # Get all time slots from DB
        ts = TimeSettings.objects.get(department=department)
        days_order = [d.name for d in ts.selected_days.all()]
        slots_order = [f"{slot.start_time.strftime('%H:%M')} – {slot.end_time.strftime('%H:%M')}" for slot in ts.selected_slots.all()]

        # Get unique faculties from timetable
        faculty_set = set()
        for entry in entries:
            faculty_set.add(entry.faculty.full_name)
        faculty_list = sorted(faculty_set)

        # Build nested dict grid[slot][day][faculty] = "Batch (Sub) [Room/Lab]"
        grid = defaultdict(lambda: defaultdict(dict))
        for entry in entries:
            time = entry.time.replace("-", "–").replace(" ", "")
            slot_match = None
            for slot in slots_order:
                if time == slot.replace(" ", ""):
                    slot_match = slot
                    break
            if not slot_match:
                continue
            cell_text = f"{entry.batch.name} ({entry.subject.subject_name})"
            room = entry.room.name if entry.room else (entry.lab.name if entry.lab else "")
            if room:
                cell_text += f" [{room}]"
            grid[slot_match][entry.day][entry.faculty.full_name] = cell_text

        # ==== HEADER ====
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(faculty_list))
        ws["A1"] = "L. J. INSTITUTE OF ENGINEERING & TECHNOLOGY, AHMEDABAD"
        ws["A1"].font = bold_font
        ws["A1"].alignment = center_align

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2 + len(faculty_list))
        ws["A2"] = "FACULTYWISE TIMETABLE (EVEN SEMESTER)"
        ws["A2"].font = bold_font
        ws["A2"].alignment = center_align

        # ==== COLUMN HEADERS ====
        header_row = ["Day", "Time Slot"] + faculty_list
        ws.append(header_row)
        for col_num in range(1, len(header_row) + 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = border
            cell.fill = gray_fill

        # ==== DATA ROWS ====
        current_row = 4
        for day in days_order:
            start_row = current_row
            for i, slot in enumerate(slots_order):
                row_data = ["", slot]
                for faculty in faculty_list:
                    row_data.append(grid[slot].get(day, {}).get(faculty, ""))
                ws.append(row_data)

                for col_num in range(1, len(row_data) + 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.alignment = center_align
                    cell.border = border
                current_row += 1

                # RECESS after second lecture
                if i == 1:
                    ws.append(["", "RECESS"] + [""] * len(faculty_list))
                    for col_num in range(2, 2 + len(faculty_list)):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.alignment = center_align
                        cell.font = bold_font
                        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                        cell.border = border
                    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=1 + len(faculty_list))
                    current_row += 1

            # Merge Day column
            ws.merge_cells(start_row=start_row, start_column=1, end_row=current_row - 1, end_column=1)
            ws.cell(row=start_row, column=1).value = day
            ws.cell(row=start_row, column=1).font = bold_font
            ws.cell(row=start_row, column=1).alignment = center_align
            ws.cell(row=start_row, column=1).border = border

            # Extra empty row
            ws.append([])
            current_row += 1
        total_row = ["", "TOTAL LECTURES"]
        for faculty in faculty_list:
            count = 0
            for slot in slots_order:
                for day in days_order:
                    if faculty in grid[slot].get(day, {}):
                        count += 1
            total_row.append(count)

        ws.append(total_row)
        total_row_index = current_row
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # light blue
        for col_num in range(1, len(total_row) + 1):
            cell = ws.cell(row=total_row_index, column=col_num)
            cell.alignment = center_align
            cell.font = Font(bold=True)
            cell.fill = blue_fill
            cell.border = border

        current_row += 1
        # Auto-width columns
        from openpyxl.cell.cell import MergedCell
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col_cells:
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Facultywise_Timetable_{selected_tt.name}.xlsx'
        wb.save(response)
        return response

    if request.GET.get("download") == "facultywisepdf" and selected_tt:
        from xhtml2pdf import pisa
        from django.template.loader import render_to_string
        from django.http import HttpResponse
        from io import BytesIO

        # Fetch time settings
        ts = TimeSettings.objects.get(department=department)
        days_order = [d.name for d in ts.selected_days.all()]
        slots_order = [f"{slot.start_time.strftime('%H:%M')} – {slot.end_time.strftime('%H:%M')}" for slot in ts.selected_slots.all()]

        # Get faculty list
        faculty_names = sorted(set(e.faculty.full_name for e in entries))

        # Prepare grid[faculty][day][slot]
        grid = {
            faculty: {
                day: {slot: "" for slot in slots_order}
                for day in days_order
            } for faculty in faculty_names
        }

        # Count lectures
        faculty_lecture_counts = {faculty: 0 for faculty in faculty_names}

        for entry in entries:
            faculty = entry.faculty.full_name
            day = entry.day
            time = entry.time.strip().replace("–", "–").replace("-", "–")
            matched_slot = next((s for s in slots_order if s.replace(" ", "") == time.replace(" ", "")), None)
            if matched_slot:
                subject = entry.subject.subject_name
                batch = entry.batch.name
                room = entry.room.name if entry.room else (entry.lab.name if entry.lab else "")
                grid[faculty][day][matched_slot] = f"{subject} ({batch})" + (f" ({room})" if room else "")
                faculty_lecture_counts[faculty] += 1

        # Render PDF
        html = render_to_string("core/facultywise_timetable_pdf.html", {
            "days_order": days_order,
            "slots_order": slots_order,
            "faculty_names": faculty_names,
            "grid": grid,
            "faculty_lecture_counts": faculty_lecture_counts,
        })

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f"attachment; filename=Facultywise_Timetable_{selected_tt.name}.pdf"
        pisa.CreatePDF(src=html, dest=response)
        return response

    # Download as Excel, CSV, PDF - GRID VIEW
    if request.GET.get("download") == "excel" and selected_tt:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = "TT-CLASSWISE"

        # Styles
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        fill_gray = PatternFill("solid", fgColor="DDDDDD")
        fill_recess = PatternFill("solid", fgColor="FFDD99")
        bold_black_border = Border(
            left=Side(style="medium", color="000000"),
            right=Side(style="medium", color="000000"),
            top=Side(style="medium", color="000000"),
            bottom=Side(style="medium", color="000000")
        )

        # Title Rows
        ws.merge_cells("A1:L1")
        ws["A1"] = "L. J. INSTITUTE OF ENGINEERING & TECHNOLOGY, AHMEDABAD"
        ws["A1"].font = header_font
        ws["A1"].alignment = center

        ws.merge_cells("A2:L2")
        ws["A2"] = "SECOND YEAR CE/IT-1 ENGG. DEPT. TIME TABLE FOR EVEN SEMESTER"
        ws["A2"].font = header_font
        ws["A2"].alignment = center

        # Header Row
        header_row = ["Day", "Timings"] + batches + ["", "", "", "", "Timings"]
        ws.append(header_row)
        for col in range(1, len(header_row) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = header_font
            cell.alignment = center
            cell.fill = fill_gray
            cell.border = bold_black_border

        row_counter = 4
        for day in days_order:
            start_row = row_counter
            for i, slot in enumerate(slots_order):
                row = ["", slot]
                for batch in batches:
                    row.append(grid[day][slot][batch])
                row += ["", "", "", "", slot]
                ws.append(row)

                for col in range(1, len(header_row) + 1):
                    cell = ws.cell(row=row_counter, column=col)
                    cell.alignment = center
                    cell.border = bold_black_border
                row_counter += 1

                if i == 1:
                    ws.append(["", "", "RECESS"] + [""] * (len(batches) - 1) + ["", "", "", "", ""])
                    for col in range(2, 2 + len(batches)):
                        cell = ws.cell(row=row_counter, column=col)
                        cell.alignment = center
                        cell.font = header_font
                        cell.fill = fill_recess
                        cell.border = bold_black_border
                    ws.merge_cells(
                        start_row=row_counter,
                        start_column=3,
                        end_row=row_counter,
                        end_column=2 + len(batches)
                    )
                    row_counter += 1

            ws.merge_cells(start_row=start_row, start_column=1, end_row=row_counter - 1, end_column=1)
            cell = ws.cell(row=start_row, column=1)
            cell.value = day
            cell.alignment = center
            cell.font = header_font
            cell.border = bold_black_border

            ws.append([])
            row_counter += 1

        # Auto column widths
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

        # ========== FACULTY + SUBJECT APPEND BELOW ==========

        faculty_subjects = []
        seen_pairs = set()

        for entry in entries:
            full_name = entry.faculty.full_name

            subject_name = entry.subject.subject_name
            key = (full_name, subject_name)
            if key not in seen_pairs:
                seen_pairs.add(key)
                faculty_subjects.append(key)

        ws.append([])
        ws.append([])
        row_counter += 3

        ws.append(["", "", "FACULTY INITIAL WITH FULL NAME AND THEIR TEACHING SUBJECTS"])
        ws.merge_cells(start_row=row_counter, start_column=3, end_row=row_counter, end_column=8)
        ws.cell(row=row_counter, column=3).font = header_font
        ws.cell(row=row_counter, column=3).alignment = center
        row_counter += 1

        table_header = ["Sr. No", "Name of Faculty", "Name of Subject"]
        ws.append(["", ""] + table_header)
        for i in range(3, 6):
            cell = ws.cell(row=row_counter, column=i)
            cell.font = header_font
            cell.alignment = center
            cell.fill = fill_gray
            cell.border = bold_black_border
        row_counter += 1

        for idx, (faculty, subject) in enumerate(faculty_subjects, 1):
            ws.append(["", "", idx, faculty, subject])
            for col in range(3, 6):
                cell = ws.cell(row=row_counter, column=col)
                cell.alignment = center
                cell.border = bold_black_border
            row_counter += 1

        # Serve the file as download
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Timetable_Styled_{selected_tt.name}.xlsx'
        wb.save(response)
        return response


    import pandas as pd
    if request.GET.get("download") == "csv" and selected_tt:
        data = []
        for day in days_order:
            for batch in batches:
                row = {"Day": day, "Batch": batch}
                for slot in slots_order:
                    row[slot] = grid[day][slot][batch]
                data.append(row)
        df = pd.DataFrame(data)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename=Timetable_{selected_tt.name}.csv'
        df.to_csv(response, index=False)
        return response

    from xhtml2pdf import pisa
    from django.template.loader import render_to_string
    from django.shortcuts import render, redirect
    from django.http import HttpResponse
    if request.GET.get("download") == "pdf" and selected_tt:
        

        # Faculty-subject list
        faculty_subjects = []
        seen_pairs = set()
        for entry in entries:
            full_name = entry.faculty.full_name
            subject_name = entry.subject.subject_name
            key = (full_name, subject_name)
            if key not in seen_pairs:
                seen_pairs.add(key)
                faculty_subjects.append(key)
        colspan_total = 2 + len(batches) + 1

        html = render_to_string("core/timetable_grid_pdf.html", {
            "days_order": days_order,
            "slots_order": slots_order,
            "batches": batches,
            "grid": grid,
            "faculty_subjects": faculty_subjects,
            "colspan_total": colspan_total
        })

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=Timetable_{selected_tt.name}.pdf'
        pisa.CreatePDF(html, dest=response)
        return response


    return render(request, "core/view_past_timetable.html", {
        "timetables": timetables,
        "selected_tt": selected_tt,
        "days_order": days_order,
        "slots_order": slots_order,
        "batches": batches,
        "grid": grid,
    })
    
from django.shortcuts import render, get_object_or_404, redirect
from .models import Timetable, TimetableEntry, Batch, Faculty, Room, Lab, CourseSpec, TimeSettings

from core.templatetags.custom_filters import get_item

def normalize_slot(slot):
    # Robust comparison by stripping all spaces and using dash
    return slot.replace(' ', '').replace('–', '-').replace('—', '-')

def edit_timetable(request, timetable_id):
    timetable = get_object_or_404(Timetable, id=timetable_id)
    department = timetable.department

    # Timetable meta info
    time_settings = TimeSettings.objects.get(department=department)
    days_order = [d.name for d in time_settings.selected_days.all()]
    slots_order = [f"{s.start_time.strftime('%H:%M')} – {s.end_time.strftime('%H:%M')}" for s in time_settings.selected_slots.all()]
    batches = [b.name for b in Batch.objects.filter(department=department)]
    subjects = CourseSpec.objects.filter(department=department)
    faculties = Faculty.objects.filter(department=department)
    rooms = Room.objects.filter(department=department)
    labs = Lab.objects.filter(department=department)
    subject_types = {str(s.id): s.subject_type for s in subjects}

    # Fetch timetable entries for this timetable
    entries = TimetableEntry.objects.filter(timetable=timetable)

    # Build grid
    grid = {}
    for day in days_order:
        grid[day] = {}
        for slot in slots_order:
            grid[day][slot] = {}
            for batch in batches:
                entry = next(
                    (e for e in entries if e.day == day and
                     normalize_slot(e.time) == normalize_slot(slot) and
                     e.batch.name == batch),
                    None
                )
                grid[day][slot][batch] = {
                    'subject_id': str(entry.subject.id) if entry and entry.subject else '',
                    'faculty_id': str(entry.faculty.id) if entry and entry.faculty else '',
                    'room_id': str(entry.room.id) if entry and entry.room else '',
                    'lab_id': str(entry.lab.id) if entry and entry.lab else '',
                    'entry_id': entry.id if entry else None
                }

    # POST: Save all changes
    if request.method == "POST":
        for day in days_order:
            for slot in slots_order:
                for batch in batches:
                    prefix = f"{day}_{slot}_{batch}"
                    subject_id = request.POST.get(f"subject_{day}_{slot}_{batch}")
                    faculty_id = request.POST.get(f"faculty_{day}_{slot}_{batch}")
                    room_id = request.POST.get(f"room_{day}_{slot}_{batch}")
                    lab_id = request.POST.get(f"lab_{day}_{slot}_{batch}")

                    # Skip if no subject/faculty selected
                    if not subject_id or not faculty_id:
                        continue

                    # Find or create entry for this cell
                    entry = next(
                        (e for e in entries if e.day == day and
                         normalize_slot(e.time) == normalize_slot(slot) and
                         e.batch.name == batch),
                        None
                    )
                    if entry:
                        entry.subject_id = subject_id
                        entry.faculty_id = faculty_id
                        # Use only one of room/lab based on subject type
                        subj_type = subject_types.get(subject_id, "")
                        if subj_type == "Theory":
                            entry.room_id = room_id if room_id else None
                            entry.lab = None
                        elif subj_type == "Practical":
                            entry.lab_id = lab_id if lab_id else None
                            entry.room = None
                        else:
                            entry.room = None
                            entry.lab = None
                        entry.save()
                    else:
                        subj_type = subject_types.get(subject_id, "")
                        entry = TimetableEntry(
                            timetable=timetable,
                            department=department,
                            day=day,
                            time=slot,
                            batch=Batch.objects.get(department=department, name=batch),
                            subject=CourseSpec.objects.get(pk=subject_id),
                            faculty=Faculty.objects.get(pk=faculty_id),
                        )
                        if subj_type == "Theory":
                            entry.room = Room.objects.get(pk=room_id) if room_id else None
                        elif subj_type == "Practical":
                            entry.lab = Lab.objects.get(pk=lab_id) if lab_id else None
                        entry.save()
        return redirect('edit_timetabe', timetable_id=timetable.id)

    return render(request, 'core/edit_timetable.html', {
        'timetable': timetable,
        'days_order': days_order,
        'slots_order': slots_order,
        'batches': batches,
        'subjects': subjects,
        'faculties': faculties,
        'rooms': rooms,
        'labs': labs,
        'subject_types': subject_types,
        'grid': grid,
    })
def get_blocked_faculty_slots(department):
    from .models import FacultyBlock
    blocks = FacultyBlock.objects.filter(department=department)
    result = {}
    for b in blocks:
        fac = b.faculty.short_name
        result.setdefault(fac, set())
        for item in b.blocked_slots:
            result[fac].add((item["day"], item["slot"]))
    return result


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from core.models import Faculty, Department, TimeSettings, Room, Lab
from .models import FacultyBlock, RoomBlock, LabBlock
import json

@login_required
def manage_faculty_blocks(request):
    department = get_object_or_404(Department, user=request.user)
    faculties = Faculty.objects.filter(department=department)
    settings = get_object_or_404(TimeSettings, department=department)

    days = [d.name for d in settings.selected_days.all()]
    slots = [f"{s.start_time.strftime('%H:%M')}–{s.end_time.strftime('%H:%M')}" for s in settings.selected_slots.all()]
    all_day_slots = [f"{day} | {slot}" for day in days for slot in slots]

    # Add new block
    if request.method == "POST":
        faculty_id = request.POST.get("faculty")
        selected_slots = request.POST.getlist("slots")
        if faculty_id and selected_slots:
            faculty = get_object_or_404(Faculty, id=faculty_id, department=department)
            blocked = []
            for entry in selected_slots:
                try:
                    day, slot = entry.split("|")
                    blocked.append({"day": day.strip(), "slot": slot.strip()})
                except:
                    continue
            FacultyBlock.objects.create(faculty=faculty, department=department, blocked_slots=blocked)
            return redirect("manage_faculty_blocks")

    # Delete block
    if request.GET.get("delete"):
        FacultyBlock.objects.filter(id=request.GET.get("delete")).delete()
        return redirect("manage_faculty_blocks")

    blocks = FacultyBlock.objects.filter(department=department)

    return render(request, "core/manage_faculty_blocks.html", {
        "faculties": faculties,
        "all_day_slots": all_day_slots,
        "blocks": blocks,
    })


@login_required
def manage_room_blocks(request):
    department = get_object_or_404(Department, user=request.user)
    rooms = Room.objects.filter(department=department)
    settings = get_object_or_404(TimeSettings, department=department)

    days = [d.name for d in settings.selected_days.all()]
    slots = [f"{s.start_time.strftime('%H:%M')}–{s.end_time.strftime('%H:%M')}" for s in settings.selected_slots.all()]
    all_day_slots = [f"{day} | {slot}" for day in days for slot in slots]

    # Add new block
    if request.method == "POST":
        room_id = request.POST.get("room")
        selected_slots = request.POST.getlist("slots")
        if room_id and selected_slots:
            room = get_object_or_404(Room, id=room_id, department=department)
            blocked = []
            for entry in selected_slots:
                try:
                    day, slot = entry.split("|")
                    blocked.append({"day": day.strip(), "slot": slot.strip()})
                except:
                    continue
            RoomBlock.objects.create(room=room, department=department, blocked_slots=blocked)
            return redirect("manage_room_blocks")

    # Delete block
    if request.GET.get("delete"):
        RoomBlock.objects.filter(id=request.GET.get("delete")).delete()
        return redirect("manage_room_blocks")

    blocks = RoomBlock.objects.filter(department=department)

    return render(request, "core/manage_room_blocks.html", {
        "rooms": rooms,
        "all_day_slots": all_day_slots,
        "blocks": blocks,
    })


@login_required
def manage_lab_blocks(request):
    department = get_object_or_404(Department, user=request.user)
    labs = Lab.objects.filter(department=department)
    settings = get_object_or_404(TimeSettings, department=department)

    days = [d.name for d in settings.selected_days.all()]
    slots = [f"{s.start_time.strftime('%H:%M')}–{s.end_time.strftime('%H:%M')}" for s in settings.selected_slots.all()]
    all_day_slots = [f"{day} | {slot}" for day in days for slot in slots]

    # Add new block
    if request.method == "POST":
        lab_id = request.POST.get("lab")
        selected_slots = request.POST.getlist("slots")
        if lab_id and selected_slots:
            lab = get_object_or_404(Lab, id=lab_id, department=department)
            blocked = []
            for entry in selected_slots:
                try:
                    day, slot = entry.split("|")
                    blocked.append({"day": day.strip(), "slot": slot.strip()})
                except:
                    continue
            LabBlock.objects.create(lab=lab, department=department, blocked_slots=blocked)
            return redirect("manage_lab_blocks")

    # Delete block
    if request.GET.get("delete"):
        LabBlock.objects.filter(id=request.GET.get("delete")).delete()
        return redirect("manage_lab_blocks")

    blocks = LabBlock.objects.filter(department=department)

    return render(request, "core/manage_lab_blocks.html", {
        "labs": labs,
        "all_day_slots": all_day_slots,
        "blocks": blocks,
    })

# core/views.py

from django.shortcuts import render, get_object_or_404
# from .forms import UploadExcelForm
from .parser import parse_faculty_workload_excel
from .models import College, Department, Faculty, Batch, CourseSpec, CourseAssignment
import json

def excel_upload(request):
    context = {}

    # 1. Get logged-in user’s department and college  
    # Assuming Department has a OneToOne or ForeignKey 'user' to auth.User
    department = get_object_or_404(Department, user=request.user)

    # If college field is required, ensure it exists and fix college name "LJIET"
    # You can optionally verify or get/create college here:
    college, _ = College.objects.get_or_create(name="LJIET")

    # Ensure department has the correct college (optional, if not set)
    if department.college != college:
        department.college = college
        department.save()

    # Now proceed with request processing

    if request.method == 'POST':
        if 'save_assignments' in request.POST:
            assignments = json.loads(request.POST.get('assignments_json', '[]'))
            # Use logged-in user's department directly (already obtained)
            for assgn in assignments:
                faculty_obj, _ = Faculty.objects.get_or_create(
                    department=department,
                    short_name=assgn['faculty_short'],
                    defaults={'full_name': assgn['faculty_short'], 'default_load': 0}
                )
                batch_obj, _ = Batch.objects.get_or_create(
                    department=department,
                    name=assgn['batch'],
                )
                course_spec_obj, _ = CourseSpec.objects.get_or_create(
                    department=department,
                    subject_name=assgn['subject'],
                    subject_type=assgn['type'],
                    defaults={'total_hours': assgn['hours']}
                )
                CourseAssignment.objects.get_or_create(
                    department=department,
                    subject=course_spec_obj,
                    faculty=faculty_obj,
                    batch=batch_obj,
                    defaults={'hours': assgn['hours'], 'room_or_lab': assgn['room_lab']}
                )
            context['success'] = True
            return render(request, 'core/excel_upload.html', context)

        else:
            form = UploadExcelForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    assignments = parse_faculty_workload_excel(form.cleaned_data['excel_file'])
                except Exception as e:
                    context['error'] = f"Excel parsing error: {str(e)}"
                    context['form'] = form
                    return render(request, 'core/excel_upload.html', context)
                if assignments:
                    context['assignments'] = assignments
                    context['assignments_json'] = json.dumps(assignments)
                else:
                    context['error'] = "No valid assignments found in file!"
                context['form'] = form
                return render(request, 'core/excel_upload.html', context)
            else:
                context['error'] = "Invalid form submission."
                context['form'] = form
                return render(request, 'core/excel_upload.html', context)

    else:
        form = UploadExcelForm()
        context['form'] = form
    return render(request, 'core/excel_upload.html', context)
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import Department, BatchRoomLabMapping
# from .forms import BatchRoomLabSimpleMappingForm

from django.contrib import messages

@login_required
def batch_room_lab_simple_mapping_view(request):
    department = get_object_or_404(Department, user=request.user)
    mappings = BatchRoomLabMapping.objects.filter(department=department).select_related('batch', 'room', 'lab').order_by('batch__name')

    if request.method == 'POST':

        # Handle Delete request
        if 'delete_mapping_id' in request.POST:
            mapping_id = request.POST.get('delete_mapping_id')
            try:
                mapping_to_delete = BatchRoomLabMapping.objects.get(id=mapping_id, department=department)
                mapping_to_delete.delete()
                messages.success(request, "Mapping deleted successfully.")
            except BatchRoomLabMapping.DoesNotExist:
                messages.error(request, "Mapping not found or unauthorized.")
            return redirect('batch_room_lab_simple_mapping')  # Ensure you set this name in urls.py

        # Handle regular form submission (Add/Update Mapping)
        form = BatchRoomLabSimpleMappingForm(request.POST, department=department)
        if form.is_valid():
            batch = form.cleaned_data['batch']
            room = form.cleaned_data['room']
            lab = form.cleaned_data['lab']

            mapping, created = BatchRoomLabMapping.objects.get_or_create(
                department=department,
                batch=batch,
                defaults={'room': room, 'lab': lab}
            )
            if not created:
                mapping.room = room
                mapping.lab = lab
                mapping.save()

            messages.success(request, "Mapping saved successfully.")
            return redirect('batch_room_lab_simple_mapping')

    else:
        form = BatchRoomLabSimpleMappingForm(department=department)

    return render(request, 'core/batch_room_lab_simple_mapping.html', {
        'form': form,
        'mappings': mappings,
    })
# core/views.py
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
# from core.forms import UploadExcelTimetableForm
from core.models import (
    Timetable, TimetableEntry, Faculty, CourseSpec, Batch,
    Room, Lab, Department, CourseAssignment, TimeSettings
)


from collections import defaultdict

import openpyxl
from collections import defaultdict
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
# from core.forms import UploadExcelTimetableForm
from core.models import (
    Timetable, TimetableEntry, Faculty, CourseSpec, Batch,
    Room, Lab, Department, CourseAssignment, TimeSettings
)

import openpyxl
from collections import defaultdict
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
# from core.forms import UploadExcelTimetableForm
from core.models import (
    Timetable, TimetableEntry, Faculty, CourseSpec, Batch,
    Room, Lab, Department, CourseAssignment, TimeSettings
)

@login_required
def upload_excel_timetable_view(request):
    department = get_object_or_404(Department, user=request.user)

    if request.method == "POST":
        form = UploadExcelTimetableForm(request.POST, request.FILES)
        if form.is_valid():
            name = form.cleaned_data["name"]
            file = form.cleaned_data["file"]

            # ==== 1. Get TimeSettings for normalization ====
            try:
                ts = TimeSettings.objects.get(department=department)
            except TimeSettings.DoesNotExist:
                messages.error(request, "TimeSettings not configured for your department.")
                return redirect("upload_excel_timetable")

            valid_days = {d.name.lower()[:3]: d.name for d in ts.selected_days.all()}
            valid_slots = [
                f"{slot.start_time.strftime('%H:%M')} – {slot.end_time.strftime('%H:%M')}"
                for slot in ts.selected_slots.all()
            ]

            def normalize_day(excel_day):
                """Map abbreviations like 'Mon' to full day names from TimeSettings"""
                short = excel_day.strip().lower()[:3] if excel_day else ""
                return valid_days.get(short, excel_day.strip().capitalize())

            def normalize_time_slot(excel_slot):
                """Convert Excel time text to match TimeSettings slot strings exactly"""
                ex_time = str(excel_slot).replace("to", "–").replace("-", "–").replace(" ", "")
                for slot in valid_slots:
                    if ex_time == slot.replace(" ", ""):
                        return slot
                return str(excel_slot).strip()

            def parse_fac_sub_room(cell_value):
                """
                Parse timetable cell format to extract faculty, subject, room/lab name, and subject_type
                Supports both formats:
                1. New format: DE (UMS) (301) => faculty=UMS, subject=DE, room=301, type=Theory
                2. New format: FSD-1 (DKU) (408-A) (Lab) => faculty=DKU, subject=FSD-1, lab=408-A, type=Practical
                3. Old format: PHA-FSD_2-408-A(L) => faculty=PHA, subject=FSD_2, lab=408-A, type=Practical
                4. Old format: PSK-DM-310 => faculty=PSK, subject=DM, room=310, type=Theory
                """
                if not cell_value or str(cell_value).strip() == "":
                    return None, None, None, None
                
                text = str(cell_value).strip()
                
                # Try new format first: DE (UMS) (301) or FSD-1 (DKU) (408-A) (Lab)
                import re
                
                # Pattern for new format: Subject (Faculty) (Room/Lab) (Lab|L)?
                new_pattern = r'^(.+?)\s*\(([^)]+)\)\s*\(([^)]+)\)(?:\s*\((?:Lab|L)\))?$'
                match = re.match(new_pattern, text)
                
                if match:
                    subject = match.group(1).strip()
                    faculty = match.group(2).strip()
                    room_lab = match.group(3).strip()
                    
                    # Check if it's a lab by looking for (Lab) or (L) suffix
                    is_lab = text.endswith('(Lab)') or text.endswith('(L)')
                    
                    if is_lab:
                        return faculty, subject, room_lab, "Practical"
                    else:
                        return faculty, subject, room_lab, "Theory"
                
                # Fallback to old format: PHA-FSD_2-408-A(L) or PSK-DM-310
                parts = text.split("-")
                if len(parts) >= 3:
                    faculty = parts[0].strip()
                    subject = parts[1].strip()
                    after = "-".join(parts[2:]).strip()
                    if "(L" in after or "(l" in after:
                        lab_name = after.split("(")[0].strip()
                        return faculty, subject, lab_name, "Practical"
                    else:
                        room_name = after.strip()
                        return faculty, subject, room_name, "Theory"
                
                return None, None, None, None

            # ==== 2. Load Excel sheet ====
            wb = openpyxl.load_workbook(file)
            try:
                sheet = wb["TT-CLASSWISE"]
            except KeyError:
                messages.error(request, "Sheet 'TT-CLASSWISE' not found in Excel.")
                return redirect("upload_excel_timetable")

            # ==== 3. Detect Batch columns ====
            import re

            header_row = next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))
            batches = []

            batch_pattern = re.compile(r"^[A-Za-z]+\d+$")  # e.g. A1, B10, abc2

            for col_idx, val in enumerate(header_row):
                if col_idx >= 2 and val:
                    val_str = str(val).strip()
                    if batch_pattern.match(val_str):  # only match letter(s) + digit(s)
                        batches.append((col_idx, val_str))

            # Ensure batches exist
            for _, bname in batches:
                Batch.objects.get_or_create(department=department, name=bname)


            # ==== 4. Create Timetable ====
            timetable = Timetable.objects.create(
                department=department,
                name=name,
                created_by=request.user
            )

            # Counter for total hours per Faculty–Subject–Batch
            hours_counter = defaultdict(int)

            # ==== 5. Parse the sheet row-by-row ====
            current_day = None
            for row in sheet.iter_rows(min_row=4, values_only=True):
                first_col = row[0]
                time_slot = row[1]

                # Detect and normalize new day
                if first_col and isinstance(first_col, str) and first_col.strip().upper() not in ["", "RECESS"]:
                    current_day = normalize_day(first_col)

                # Skip recess or invalid row
                if not current_day or not time_slot or (isinstance(row[2], str) and row[2].strip().upper() == "RECESS"):
                    continue

                matched_time = normalize_time_slot(time_slot)

                # Loop through each batch column
                for col_idx, batch_name in batches:
                    cell_val = row[col_idx]
                    fac, subj, rl, subj_type = parse_fac_sub_room(cell_val)
                    if not fac or not subj:
                        continue

                    faculty_obj, _ = Faculty.objects.get_or_create(
                        department=department,
                        short_name=fac,
                        defaults={"full_name": fac}
                    )

                    subject_obj, _ = CourseSpec.objects.get_or_create(
                        department=department,
                        subject_name=subj,
                        defaults={"subject_type": subj_type, "total_hours": 0}
                    )
                    if subject_obj.subject_type != subj_type:
                        subject_obj.subject_type = subj_type
                        subject_obj.save()

                    room_obj, lab_obj = None, None
                    if rl:
                        if subj_type == "Practical":
                            lab_obj, _ = Lab.objects.get_or_create(
                                department=department,
                                name=rl,
                                defaults={"capacity": 0}
                            )
                        else:
                            room_obj, _ = Room.objects.get_or_create(
                                department=department,
                                name=rl,
                                defaults={"capacity": 0}
                            )

                    batch_obj = Batch.objects.get(department=department, name=batch_name)

                    # Save TimetableEntry with normalized fields
                    TimetableEntry.objects.create(
                        timetable=timetable,
                        department=department,
                        subject=subject_obj,
                        faculty=faculty_obj,
                        batch=batch_obj,
                        day=current_day,
                        time=matched_time,
                        room=room_obj,
                        lab=lab_obj
                    )

                    hours_counter[(faculty_obj.id, subject_obj.id, batch_obj.id)] += 1

            # ==== 6. Save CourseAssignments with calculated hours ====
            for (faculty_id, subject_id, batch_id), hours in hours_counter.items():
                faculty_obj = Faculty.objects.get(id=faculty_id)
                subject_obj = CourseSpec.objects.get(id=subject_id)
                batch_obj = Batch.objects.get(id=batch_id)
                room_or_lab = "Lab" if subject_obj.subject_type == "Practical" else "Room"

                ca, created = CourseAssignment.objects.get_or_create(
                    department=department,
                    subject=subject_obj,
                    faculty=faculty_obj,
                    batch=batch_obj,
                    defaults={"hours": hours, "room_or_lab": room_or_lab}
                )
                if not created:
                    ca.hours = hours
                    ca.room_or_lab = room_or_lab
                    ca.save()

            messages.success(request, "Timetable uploaded and saved successfully!")
            return redirect("view_past_timetable")

    else:
        form = UploadExcelTimetableForm()

    return render(request, "core/upload_excel_timetable.html", {"form": form})


@login_required
def manage_timetable_type(request):
    """View for managing timetable type (1 hour vs 2 hour slots)"""
    department = get_object_or_404(Department, user=request.user)
    
    # Get or create TimetableType for this department
    timetable_type, created = TimetableType.objects.get_or_create(
        department=department,
        defaults={'slot_type': '2_hour'}
    )
    
    if request.method == 'POST':
        form = TimetableTypeForm(request.POST, instance=timetable_type)
        if form.is_valid():
            form.save()
            messages.success(request, f"Timetable type updated to {timetable_type.get_slot_type_display()}")
            return redirect('manage_timetable_type')
    else:
        form = TimetableTypeForm(instance=timetable_type)
    
    return render(request, 'core/manage_timetable_type.html', {
        'form': form,
        'timetable_type': timetable_type
    })

# Admin Module Views
@login_required
def admin_dashboard(request):
    """Admin dashboard with password protection"""
    # Check if user has admin access
    if not request.session.get('admin_access'):
        return redirect('admin_login')
    
    colleges = College.objects.all()
    departments = Department.objects.select_related('college').all()
    users = User.objects.filter(department__isnull=False).select_related('department__college')
    
    return render(request, 'core/admin_dashboard.html', {
        'colleges': colleges,
        'departments': departments,
        'users': users
    })

@login_required
def admin_login(request):
    """Admin login with password verification"""
    if request.session.get('admin_access'):
        return redirect('admin_dashboard')
    
    if request.method == 'POST':
        form = AdminPasswordForm(request.POST)
        if form.is_valid():
            password = form.cleaned_data['password']
            # You can change this password as needed
            if password == 'admin123':  # Change this to your desired admin password
                request.session['admin_access'] = True
                messages.success(request, "Admin access granted!")
                return redirect('admin_dashboard')
            else:
                messages.error(request, "Invalid admin password!")
    else:
        form = AdminPasswordForm()
    
    return render(request, 'core/admin_login.html', {'form': form})

@login_required
def admin_logout(request):
    """Admin logout"""
    request.session.pop('admin_access', None)
    messages.success(request, "Admin access revoked!")
    return redirect('admin_dashboard')

@login_required
def create_college(request):
    """Create new college"""
    if not request.session.get('admin_access'):
        return redirect('admin_login')
    
    if request.method == 'POST':
        form = CollegeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "College created successfully!")
            return redirect('admin_dashboard')
    else:
        form = CollegeForm()
    
    return render(request, 'core/create_college.html', {'form': form})

@login_required
def create_department(request):
    """Create new department"""
    if not request.session.get('admin_access'):
        return redirect('admin_login')
    
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Department created successfully!")
            return redirect('admin_dashboard')
    else:
        form = DepartmentForm()
    
    return render(request, 'core/create_department.html', {'form': form})

@login_required
def create_user(request):
    """Create new user and assign to department"""
    if not request.session.get('admin_access'):
        return redirect('admin_login')
    
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Get department from form data
            department_id = request.POST.get('department')
            if department_id:
                try:
                    department = Department.objects.get(id=department_id)
                    department.user = user
                    department.save()
                    messages.success(request, "User created and assigned to department successfully!")
                except Department.DoesNotExist:
                    messages.error(request, "Department not found!")
            else:
                messages.success(request, "User created successfully!")
            return redirect('admin_dashboard')
    else:
        form = UserForm()
    
    departments = Department.objects.filter(user__isnull=True)
    return render(request, 'core/create_user.html', {
        'form': form,
        'departments': departments
    })

@login_required
def edit_college(request, college_id):
    """Edit college"""
    if not request.session.get('admin_access'):
        return redirect('admin_login')
    
    college = get_object_or_404(College, id=college_id)
    
    if request.method == 'POST':
        form = CollegeForm(request.POST, instance=college)
        if form.is_valid():
            form.save()
            messages.success(request, "College updated successfully!")
            return redirect('admin_dashboard')
    else:
        form = CollegeForm(instance=college)
    
    return render(request, 'core/edit_college.html', {'form': form, 'college': college})

@login_required
def edit_department(request, department_id):
    """Edit department"""
    if not request.session.get('admin_access'):
        return redirect('admin_login')
    
    department = get_object_or_404(Department, id=department_id)
    
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, "Department updated successfully!")
            return redirect('admin_dashboard')
    else:
        form = DepartmentForm(instance=department)
    
    return render(request, 'core/edit_department.html', {'form': form, 'department': department})

@login_required
def delete_college(request, college_id):
    """Delete college"""
    if not request.session.get('admin_access'):
        return redirect('admin_login')
    
    college = get_object_or_404(College, id=college_id)
    college.delete()
    messages.success(request, "College deleted successfully!")
    return redirect('admin_dashboard')

@login_required
def delete_department(request, department_id):
    """Delete department"""
    if not request.session.get('admin_access'):
        return redirect('admin_login')
    
    department = get_object_or_404(Department, id=department_id)
    department.delete()
    messages.success(request, "Department deleted successfully!")
    return redirect('admin_dashboard')

@login_required
def delete_user(request, user_id):
    """Delete user"""
    if not request.session.get('admin_access'):
        return redirect('admin_login')
    
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.success(request, "User deleted successfully!")
    return redirect('admin_dashboard')


# Faculty Preferred Slots Views
@login_required
def manage_faculty_preferred_slots(request):
    """Manage faculty preferred slots with same design as blocked slots"""
    department = get_object_or_404(Department, user=request.user)
    faculties = Faculty.objects.filter(department=department)
    batches = Batch.objects.filter(department=department)
    settings = get_object_or_404(TimeSettings, department=department)

    days = [d.name for d in settings.selected_days.all()]
    slots = [f"{s.start_time.strftime('%H:%M')}–{s.end_time.strftime('%H:%M')}" for s in settings.selected_slots.all()]
    all_day_slots = [f"{day} | {slot}" for day in days for slot in slots]

    # Add new preferred slots
    if request.method == "POST":
        faculty_id = request.POST.get("faculty")
        batch_id = request.POST.get("batch")
        selected_slots = request.POST.getlist("slots")
        if faculty_id and batch_id and selected_slots:
            faculty = get_object_or_404(Faculty, id=faculty_id, department=department)
            batch = get_object_or_404(Batch, id=batch_id, department=department)
            preferred = []
            for entry in selected_slots:
                try:
                    day, slot = entry.split("|")
                    preferred.append({"day": day.strip(), "slot": slot.strip()})
                except:
                    continue
            FacultyPreferredSlot.objects.create(
                faculty=faculty, 
                batch=batch, 
                department=department, 
                preferred_slots=preferred
            )
            messages.success(request, "Preferred slots added successfully!")
            return redirect("manage_faculty_preferred_slots")

    # Delete preferred slots
    if request.GET.get("delete"):
        FacultyPreferredSlot.objects.filter(id=request.GET.get("delete"), department=department).delete()
        messages.success(request, "Preferred slots deleted successfully!")
        return redirect("manage_faculty_preferred_slots")

    preferred_slots = FacultyPreferredSlot.objects.filter(department=department)

    return render(request, "core/manage_faculty_preferred_slots.html", {
        "faculties": faculties,
        "batches": batches,
        "all_day_slots": all_day_slots,
        "preferred_slots": preferred_slots,
    })

@login_required
def faculty_preferred_slots_list(request):
    """List all faculty preferred slots for the current department with same design as blocked slots"""
    department = get_object_or_404(Department, user=request.user)
    faculties = Faculty.objects.filter(department=department)
    batches = Batch.objects.filter(department=department)
    settings = get_object_or_404(TimeSettings, department=department)

    days = [d.name for d in settings.selected_days.all()]
    slots = [f"{s.start_time.strftime('%H:%M')}–{s.end_time.strftime('%H:%M')}" for s in settings.selected_slots.all()]
    all_day_slots = [f"{day} | {slot}" for day in days for slot in slots]

    # Add new preferred slots
    if request.method == "POST":
        faculty_id = request.POST.get("faculty")
        batch_id = request.POST.get("batch")
        selected_slots = request.POST.getlist("slots")
        if faculty_id and batch_id and selected_slots:
            faculty = get_object_or_404(Faculty, id=faculty_id, department=department)
            batch = get_object_or_404(Batch, id=batch_id, department=department)
            preferred = []
            for entry in selected_slots:
                try:
                    day, slot = entry.split("|")
                    preferred.append({"day": day.strip(), "slot": slot.strip()})
                except:
                    continue
            FacultyPreferredSlot.objects.create(
                faculty=faculty, 
                batch=batch, 
                department=department, 
                preferred_slots=preferred
            )
            messages.success(request, "Preferred slots added successfully!")
            return redirect("faculty_preferred_slots_list")

    # Delete preferred slots
    if request.GET.get("delete"):
        FacultyPreferredSlot.objects.filter(id=request.GET.get("delete"), department=department).delete()
        messages.success(request, "Preferred slots deleted successfully!")
        return redirect("faculty_preferred_slots_list")

    preferred_slots = FacultyPreferredSlot.objects.filter(department=department)
    
    return render(request, 'core/faculty_preferred_slots_list.html', {
        'faculties': faculties,
        'batches': batches,
        'all_day_slots': all_day_slots,
        'preferred_slots': preferred_slots,
        'department': department
    })


@login_required
def add_faculty_preferred_slots(request):
    """Add new faculty preferred slots"""
    department = get_object_or_404(Department, id=request.session['selected_department_id'])
    
    if request.method == 'POST':
        form = FacultyPreferredSlotForm(request.POST, department=department)
        if form.is_valid():
            preferred_slot = form.save(commit=False)
            preferred_slot.department = department
            preferred_slot.save()
            messages.success(request, "Faculty preferred slots added successfully!")
            return redirect('faculty_preferred_slots_list')
    else:
        form = FacultyPreferredSlotForm(department=department)
    
    return render(request, 'core/add_faculty_preferred_slots.html', {
        'form': form,
        'department': department
    })


@login_required
def edit_faculty_preferred_slots(request, pk):
    """Edit faculty preferred slots"""
    department = get_object_or_404(Department, id=request.session['selected_department_id'])
    preferred_slot = get_object_or_404(FacultyPreferredSlot, pk=pk, department=department)
    
    if request.method == 'POST':
        form = FacultyPreferredSlotForm(request.POST, instance=preferred_slot, department=department)
        if form.is_valid():
            form.save()
            messages.success(request, "Faculty preferred slots updated successfully!")
            return redirect('faculty_preferred_slots_list')
    else:
        form = FacultyPreferredSlotForm(instance=preferred_slot, department=department)
    
    return render(request, 'core/edit_faculty_preferred_slots.html', {
        'form': form,
        'preferred_slot': preferred_slot,
        'department': department
    })


@login_required
def delete_faculty_preferred_slots(request, pk):
    """Delete faculty preferred slots"""
    department = get_object_or_404(Department, id=request.session['selected_department_id'])
    preferred_slot = get_object_or_404(FacultyPreferredSlot, pk=pk, department=department)
    
    if request.method == 'POST':
        preferred_slot.delete()
        messages.success(request, "Faculty preferred slots deleted successfully!")
        return redirect('faculty_preferred_slots_list')
    
    return render(request, 'core/delete_faculty_preferred_slots.html', {
        'preferred_slot': preferred_slot,
        'department': department
    })


def get_preferred_faculty_slots(department):
    """
    Returns a dictionary mapping (faculty_short_name, batch_name) to a set of (day, slot)
    for all preferred slots for that faculty-batch combination.
    """
    preferred_slots = FacultyPreferredSlot.objects.filter(department=department)
    result = {}
    for pref in preferred_slots:
        key = (pref.faculty.short_name, pref.batch.name)
        try:
            slot_list = pref.preferred_slots
            # Convert list of dicts to set of tuples
            result[key] = set((item["day"], item["slot"]) for item in slot_list)
            print(f"DEBUG: Preferred slots for {key}: {result[key]}")
        except Exception as e:
            print(f"DEBUG: Error processing preferred slots for {key}: {e}")
            continue
    return result


# Excel Leave Management Views
from .excel_leave_views import (
    excel_leave_management_view,
    excel_faculty_day_selection_view,
    excel_leave_details_view,
    excel_generate_temporary_timetable_view,
    clear_excel_session_view,
    download_temporary_timetable_pdf
)

