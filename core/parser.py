def parse_faculty_workload_excel(file):
    import openpyxl
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    # Read headers
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h).strip() if h else '' for h in headers]

    # Find positions
    subj_idx = headers.index('subject')
    hours_idx = headers.index('HOURS')
    type_idx = headers.index('TYPE')

    # Batch columns are all between subject and HOURS (exclusive)
    batch_start = subj_idx + 1
    batch_end = hours_idx  # exclusive

    batch_names = headers[batch_start:batch_end]

    assignments = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        subject = row[subj_idx]
        hours_per_batch = row[hours_idx]
        type_cell = row[type_idx]
        subject_type = 'Theory' if str(type_cell).strip().upper() == 'ROOM' else 'Practical'

        for i, batch in enumerate(batch_names):
            batch_idx = batch_start + i
            faculty_short = row[batch_idx] if batch_idx < len(row) else None
            if faculty_short and str(faculty_short).strip().upper() not in ('', 'NONE'):
                assignments.append({
                    'subject': subject,
                    'batch': str(batch).strip(),
                    'faculty_short': str(faculty_short).strip().upper(),
                    'hours': hours_per_batch,
                    'type': subject_type,
                    'room_lab': 'Room' if subject_type == 'Theory' else 'Lab'
                })
    return assignments
